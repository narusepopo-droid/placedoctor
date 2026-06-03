import sys
import io
import os

# 한글 깨짐 완전 방지 — Python UTF-8 모드 강제 활성화
os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("PYTHONIOENCODING", "utf-8:replace")

try:
    if hasattr(sys.stdout, 'buffer') and not isinstance(sys.stdout, io.TextIOWrapper):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
    elif hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass
try:
    if hasattr(sys.stderr, 'buffer') and not isinstance(sys.stderr, io.TextIOWrapper):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)
    elif hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

import asyncio
import re
import unicodedata
import urllib.parse
import urllib.request
import os
import json
import random
import subprocess
import threading
import gc
import socket
import traceback
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, messagebox, scrolledtext
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl
from openpyxl import Workbook

socket.setdefaulttimeout(10)

APP_VERSION = "8.33"  # 버전 정보의 단일 출처 — firebase_auth.CURRENT_VERSION에 덮어씌워짐

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

try:
    import sys as _sys, os as _os
    _self_dir = (_os.path.dirname(_os.path.abspath(__file__))
                 if '__file__' in vars() else _os.path.dirname(_os.path.abspath(_sys.argv[0])))
    if _self_dir not in _sys.path:
        _sys.path.insert(0, _self_dir)
    import firebase_auth as _fb_auth
    from firebase_auth import load_firebase_auth, LoginWindow, AdminTab, check_and_update, try_auto_login, needs_update
    _fb_auth.CURRENT_VERSION = APP_VERSION  # naver_tracker.py가 버전의 단일 출처
    FIREBASE_AVAILABLE = True
except Exception:
    FIREBASE_AVAILABLE = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "gui_config.json")

# 분석기록은 재설치해도 안 사라지도록 AppData에 저장
DATA_DIR = os.path.join(os.environ.get("APPDATA", BASE_DIR), "PlaceMasterPRO")
os.makedirs(DATA_DIR, exist_ok=True)
HISTORY_FILE = os.path.join(DATA_DIR, "analysis_history.json")

# 구버전(프로그램 폴더)에 기록 파일이 있으면 AppData로 자동 이전
_old_history = os.path.join(BASE_DIR, "analysis_history.json")
if os.path.exists(_old_history) and not os.path.exists(HISTORY_FILE):
    try:
        import shutil
        shutil.move(_old_history, HISTORY_FILE)
    except Exception:
        pass

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_history(history):
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def merge_history(local_hist: dict, remote_hist: dict) -> dict:
    """로컬 + Firebase 기록 병합. 같은 날짜+모드는 로컬 우선."""
    merged = {k: list(v) for k, v in local_hist.items()}
    for store_name, remote_entries in remote_hist.items():
        if store_name not in merged:
            merged[store_name] = list(remote_entries)
        else:
            local_keys = {(e['date'], e['mode']) for e in merged[store_name]}
            for entry in remote_entries:
                if (entry['date'], entry['mode']) not in local_keys:
                    merged[store_name].append(entry)
            merged[store_name].sort(key=lambda e: e['date'])
            merged[store_name] = merged[store_name][-60:]
    return merged
EXCEL_FILE = os.path.join(BASE_DIR, "통합_마케팅_리포트.xlsx")

def normalize_config(config_data=None):
    """설정 파일이 비어 있거나 예전 버전 구조여도 안전하게 보정합니다."""
    if not isinstance(config_data, dict):
        config_data = {}
    if not isinstance(config_data.get("stores"), list):
        config_data["stores"] = []
    for s in config_data["stores"]:
        if not isinstance(s.get("custom_keywords"), list):
            s["custom_keywords"] = []
    if "safe_mode" not in config_data:
        config_data["safe_mode"] = True
    if not config_data.get("kakao_style"):
        config_data["kakao_style"] = "이모지 Trendy형"
    return config_data

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return normalize_config(json.load(f))
        except Exception:
            pass
    return normalize_config({"safe_mode": True, "kakao_style": "이모지 Trendy형", "stores": []})

def save_config(config_data):
    os.makedirs(BASE_DIR, exist_ok=True)
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(normalize_config(config_data), f, ensure_ascii=False, indent=4)

async def get_place_details_failsafe(page, url, log_func):
    try:
        # 1단계: 원본 URL로 place_id 추출
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(1000)

        p_id = None
        match = re.search(r'\d{8,11}', page.url)
        if match: p_id = match.group(0)
        if not p_id: return None, "", "", [], [], "", []

        # 2단계: map.naver.com 진입 후 keywordList 추출 (메인 페이지 JSON에 포함)
        map_url = f"https://map.naver.com/p/entry/place/{p_id}"
        await page.goto(map_url, wait_until="domcontentloaded", timeout=25000)
        await page.wait_for_timeout(3000)

        _kl_js = '''() => {
            const fullHtml = document.documentElement.innerHTML;
            const pats = [
                /"keywordList"\\s*:\\s*(\\[[^\\]]{2,2000}\\])/,
                /keywordList["']?\\s*:\\s*(\\[[^\\]]{2,2000}\\])/
            ];
            for (const pat of pats) {
                const m = fullHtml.match(pat);
                if (m) { try { const r = JSON.parse(m[1]); if (r && r.length) return r; } catch(e) {} }
            }
            return [];
        }'''
        # keywordList는 메인 페이지 script JSON에 있음 — 최대 2회 시도 (페이지 늦게 로드 대비)
        keyword_list = await page.evaluate(_kl_js)
        if not keyword_list:
            await page.wait_for_timeout(2000)
            keyword_list = await page.evaluate(_kl_js)

        # 3단계: pcmap iframe → 주소/카테고리 추출
        # 타임아웃 시 place_id·keywordList는 이미 확보된 상태 유지 (별도 try)
        iframe_src = await page.evaluate('''() => {
            for (const f of document.querySelectorAll("iframe")) {
                const src = f.src || "";
                if (src.includes("pcmap.place.naver.com") || src.includes("place.map.naver.com")) return src;
            }
            return "";
        }''')

        if iframe_src:
            try:
                await page.goto(iframe_src, wait_until="domcontentloaded", timeout=25000)
                await page.wait_for_timeout(3000)
                try:
                    await page.wait_for_selector(".pz7wy, .PkgBl, .LDgIH, .IH7VW, [class*='addr']", timeout=5000)
                except Exception:
                    pass
                # iframe에서도 keywordList 재시도 (메인 페이지에서 못 잡은 경우)
                if not keyword_list:
                    keyword_list = await page.evaluate('''() => {
                        const fullHtml = document.documentElement.innerHTML;
                        const m = fullHtml.match(/"keywordList"\\s*:\\s*(\\[[^\\]]{2,400}\\])/);
                        if (m) { try { const r = JSON.parse(m[1]); if (r && r.length) return r; } catch(e) {} }
                        return [];
                    }''')
            except Exception as iframe_err:
                log_func(f"    ㄴ ⚠️ iframe 로딩 지연 (주소 추출 건너뜀): {type(iframe_err).__name__}")

        details = await page.evaluate('''() => {
            const REGIONS = ["서울","경기","인천","부산","대구","광주","대전","울산","세종","강원","충북","충남","전북","전남","경북","경남","제주"];
            // 주소 추출 — CSS 셀렉터 우선, script JSON 폴백, span/div 광역 스캔
            const addrSels = [".pz7wy",".PkgBl",".LDgIH",".IH7VW","[class*='addr_area']","[class*='address']","address"];
            let addrEl = null;
            for (const sel of addrSels) {
                const el = document.querySelector(sel);
                if (el && (el.innerText||"").trim().length > 5) { addrEl = el; break; }
            }
            let addr = addrEl ? (addrEl.innerText || "").split("\\n")[0].trim() : "";

            // script JSON에서 roadAddress/address 키 추출
            if (!addr) {
                for (const s of document.querySelectorAll("script:not([src])")) {
                    const t = s.textContent || "";
                    const m = t.match(/"(?:roadAddress|address|jibunAddress)"\\s*:\\s*"([^"]{5,80})"/);
                    if (m && REGIONS.some(r => m[1].includes(r))) { addr = m[1]; break; }
                }
            }
            // 광역 span/a/div 스캔
            if (!addr) {
                for (const el of document.querySelectorAll("span, a, div, p")) {
                    const t = (el.innerText || "").trim().split("\\n")[0].trim();
                    if (t.length > 5 && t.length < 80 && REGIONS.some(r => t.includes(r)) && (t.includes("구") || t.includes("동") || t.includes("로") || t.includes("길"))) {
                        addr = t; break;
                    }
                }
            }
            const cat = document.querySelector(".DJJvD, .lnJFt")?.innerText?.trim() || "";

            // 동 추출
            let dongFound = "";
            if (addrEl) {
                const m = (addrEl.innerText || "").match(/([가-힣]{2,6}동)(?=[\\s\\d\\-·,]|$)/);
                if (m) dongFound = m[1];
            }
            if (!dongFound && addr) {
                const m = addr.match(/([가-힣]{2,6}동)/);
                if (m) dongFound = m[1];
            }
            if (!dongFound) {
                const bodyText = document.body ? (document.body.innerText || "") : "";
                const m = bodyText.match(/[시군구]\\s+([가-힣]{2,6}동)/);
                if (m && m[1] && m[1].length >= 3) dongFound = m[1];
            }
            if (!dongFound) {
                for (const s of document.querySelectorAll("script:not([src])")) {
                    const t = s.textContent || "";
                    const m = t.match(/"(?:legalDong|dong|eupMyeonDong|address)"\\s*:\\s*"([가-힣]{2,6}동)/);
                    if (m) { dongFound = m[1]; break; }
                }
            }

            // 역 추출 — 출구 패턴 우선, 없으면 도보/차량 거리 패턴
            let nearbyStation = "";
            const bodyText2 = document.body ? (document.body.innerText || "") : "";
            const stM = bodyText2.match(/([가-힣]{2,8}역)\s*(?:\d+번\s*출구|에서\s*(?:도보|차로|차량|\d))/);
            if (stM) nearbyStation = stM[1];
            if (!nearbyStation) {
                const stM2 = bodyText2.match(/([가-힣]{2,8}역)\s*(?:도보|방향|하차|인근|근처)/);
                if (stM2) nearbyStation = stM2[1];
            }

            return { address: addr, category: cat, nearbyDong: dongFound, nearbyStation };
        }''')

        menu_items = await page.evaluate('''() => {
            const menus = Array.from(document.querySelectorAll(".lPzOq.VXIyT, .Sqg65, .A_cdD, .y20fl"));
            return menus.slice(0, 5).map(el => el.innerText).filter(text => text && text.length > 1);
        }''')

        official_keywords = await page.evaluate('''() => {
            const tags = Array.from(document.querySelectorAll(".PR2aT .P1zUJ, .P1zUJ, .hB45E, .place_bluelink"));
            return tags.map(el => el.innerText.replace(/#/g, '')).filter(text => text && text.length > 1);
        }''')

        _ADDR_REGIONS = ["서울","경기","인천","부산","대구","광주","대전","울산","세종","강원","충북","충남","전북","전남","경북","경남","제주"]
        _ADDR_SUFFIXES = ["구","동","로","길","읍","면","리"]
        def _valid_addr(a):
            """한국 주소 형식 검증 — 광역시도 + 행정단위 포함 여부"""
            return bool(a) and any(r in a for r in _ADDR_REGIONS) and any(s in a for s in _ADDR_SUFFIXES)

        address_full = details["address"]
        # 한국 주소 형식이 아니면 쓰레기값으로 판단해 비워둠 → 모바일 폴백 트리거
        if not _valid_addr(address_full):
            address_full = ""
        nearby_dong = details.get("nearbyDong", "")
        nearby_station = details.get("nearbyStation", "")

        # nearby_dong 폴백: jibunAddress에서 동 이름 추출 (도로명 주소에는 동이 없으므로)
        if not nearby_dong:
            jibun = await page.evaluate('''() => {
                for (const s of document.querySelectorAll("script:not([src])")) {
                    const t = s.textContent || "";
                    // jibunAddress: "서울특별시 노원구 상계동 723-2" 형태
                    const m1 = t.match(/"jibunAddress"\\s*:\\s*"([^"]+)"/);
                    if (m1) return m1[1];
                    // legalDong: "상계동" 직접 명시
                    const m2 = t.match(/"legalDong"\\s*:\\s*"([가-힣]{2,6}동)"/);
                    if (m2) return m2[1];
                    // 구군 바로 다음에 동 이름 (주소 컨텍스트)
                    const m3 = t.match(/[구군]\\s+([가-힣]{2,6}동)/);
                    if (m3) return m3[1];
                }
                return "";
            }''')
            if jibun:
                m = re.search(r'([가-힣]{2,6}동)', jibun)
                if m:
                    nearby_dong = m.group(1)

        # 주소 추출 실패 시 → 모바일 URL 폴백
        if not address_full and p_id:
            try:
                await page.goto(f"https://m.place.naver.com/place/{p_id}/home",
                                wait_until="domcontentloaded", timeout=12000)
                await page.wait_for_timeout(2000)
                mob = await page.evaluate('''() => {
                    const REGIONS = ["서울","경기","인천","부산","대구","광주","대전","울산","세종","강원","충북","충남","전북","전남","경북","경남","제주"];
                    // script JSON 우선
                    for (const s of document.querySelectorAll("script:not([src])")) {
                        const t = s.textContent || "";
                        const m = t.match(/"(?:roadAddress|address|jibunAddress)"\\s*:\\s*"([^"]{5,80})"/);
                        if (m && REGIONS.some(r => m[1].includes(r))) return { addr: m[1], dong: "", station: "" };
                    }
                    // DOM 스캔
                    for (const el of document.querySelectorAll("span, div, p, address")) {
                        const t = (el.innerText || "").trim().split("\\n")[0].trim();
                        if (t.length > 5 && t.length < 80 && REGIONS.some(r => t.includes(r)) && (t.includes("구") || t.includes("동") || t.includes("로") || t.includes("길"))) {
                            const dm = t.match(/([가-힣]{2,6}동)/);
                            const sm = (document.body.innerText||"").match(/([가-힣]{2,8}역)\\s*(?:\\d+번\\s*출구|에서\\s*(?:도보|차로|차량))/);
                            return { addr: t, dong: dm ? dm[1] : "", station: sm ? sm[1] : "" };
                        }
                    }
                    return { addr: "", dong: "", station: "" };
                }''')
                if mob.get("addr"):
                    address_full = mob["addr"]
                    if not nearby_dong: nearby_dong = mob.get("dong", "")
                    if not nearby_station: nearby_station = mob.get("station", "")
            except Exception:
                pass

        log_func(f"    ㄴ 📍 주소: {address_full} | 동: {nearby_dong or '(없음)'} | 역: {nearby_station or '(없음)'}")
        if keyword_list:
            log_func(f"    ㄴ 🔑 대표 키워드: {', '.join(keyword_list)}")
        # nearbyDong을 주소에 항상 병합 — DOM에서 추출된 정확한 동 이름이므로 서울/경기 포함 전국 적용
        if nearby_dong and nearby_dong not in address_full:
            address_full = address_full + " " + nearby_dong
        return p_id, address_full, details["category"], menu_items, official_keywords, nearby_station, keyword_list
    except Exception as e:
        log_func(f"    ㄴ ⚠️ 정보 수집 지연: {e}")
        return None, "", "", [], [], "", []

# ── 검색 의도 토큰 사전 (keywordList 분해용) ─────────────────────────────────
_INTENT_TOKENS = [
    # 캠핑/야외
    "오토캠핑장","감성캠핑장","가족캠핑장","커플캠핑장","캠핑장","글램핑장","야영장",
    "오토캠핑","감성캠핑","가족캠핑","커플캠핑","글램핑","야영","카라반","차박여행","차박","계곡캠핑","캠핑",
    # 숙박
    "계곡펜션","풀빌라","독채펜션","가족펜션","커플펜션","계곡숙박",
    "펜션","리조트","숙박","게스트하우스","민박","호텔","콘도",
    # 자연/여행
    "계곡여행","가족여행","커플여행","당일치기","1박2일",
    "계곡","강변","호수","바다","산","여행","관광","나들이","힐링","드라이브","명소","체험",
    # 음식점 — 시간대/목적
    "점심특선","저녁특선","점심맛집","저녁맛집","야식맛집","아침식사","브런치",
    "점심","저녁","아침","특선","런치","포장","테이크아웃","혼밥","혼술단체",
    # 음식점 — 메뉴별
    "돼지국밥맛집","갈비찜맛집","등갈비맛집","고기맛집",
    "돼지국밥","순대국밥","갈비찜","등갈비","돼지등","등뼈찜","해장국","순대국","국밥",
    "쭈꾸미","수육","감자탕","뼈다귀탕","도가니탕","설렁탕","곰탕","부대찌개","김치찌개","된장찌개",
    "삼겹살","돼지갈비","갈비","고기집","맛집","식당","한식당","회식","단체석","혼술","술집","야식",
    "이자카야","호프","포차","막걸리집","안주","안주맛집","소주","맥주","내돈내산",
    "족발맛집","보쌈맛집","족발","보쌈","곱창맛집","곱창","막창",
    "카페맛집","브런치카페","감성카페","루프탑카페","야경카페","정원카페","좋은카페","대형카페",
    "베이커리카페","디저트카페","카페테라스","대형베이커리카페","대형테라스카페","뷰카페","포토카페","숲속카페",
    "카페","커피","디저트","케이크","베이커리","마카롱","크로플","와플",
    "오마카세","초밥","스시","일식","돈까스","라멘","우동",
    "피자","파스타","스테이크","양식","치킨","닭갈비","닭볶음탕",
    "쌀국수","베트남음식","짬뽕","짜장면","탕수육","중식","마라탕",
    "냉면","막국수","떡볶이","분식","뷔페","무한리필","무한주류",
    # 피트니스/스포츠
    "퍼스널트레이닝","개인PT","헬스클럽","피트니스센터","다이어트PT","바디프로필",
    "재활PT","헬스장","피트니스","필라테스","요가","스포츠센터","PT","크로스핏","수영장",
    "골프","스크린골프","실내골프","골프연습장","골프레슨","그룹PT","그룹필라테스","체형교정",
    # 학원/교육 — 복합 키워드
    "영재학원","사고력학원","사고력수학","사고력교육","창의수학","영재수학","과학영재","수학경시",
    "유아수학","초등수학","중등수학","조기수학","유아교육","초등교육","영재교육","조기교육",
    "영재입시","과학고입시","영재고입시","과학학원","과학교육",
    "수학학원","영어학원","코딩학원","입시학원","피아노학원","미술학원","음악학원","태권도학원",
    "학원","교육센터","공부방","독서실","스터디카페",
    # 학원/교육 — 내신 관련
    "내신관리","내신대비","내신전문","내신수학","내신영어","내신국어","내신준비",
    # 학원/교육 — 루트 토큰 (keywordList 분해용: "사고력유아수학" → "사고력"+"유아"+"수학")
    "사고력","영재","수학","영어","유아","초등","중등","고등","입시","내신","특기","논술","과학",
    # 미용/헤어
    "미용실","헤어샵","헤어살롱","미용원","머리잘하는곳",
    "커트","펌","염색","탈색","두피케어","헤어트리트먼트","매직","셋팅펌",
    "네일샵","젤네일","네일아트","속눈썹연장","왁싱","눈썹문신","반영구",
    # 피부/뷰티
    "피부관리","피부케어","피부미용","에스테틱","에스테틱샵","피부샵","관리샵",
    "윤곽관리","윤곽마사지","여드름관리","여드름케어","여드름","모공관리","피지관리",
    "웨딩관리","웨딩케어","리프팅","탄력관리","미백관리","수분관리",
    "마사지","스파","아로마","림프마사지","왁싱관리",
    "피부과","피부과의원","레이저","보톡스","필러","성형외과",
    # 피부과 시술·기기 (써마지/울쎄라 등 오인식 방지 — location으로 잘못 분류되면 안 됨)
    "써마지","써마지FLX","울쎄라","울쎄라리프팅","슈링크","인모드","포텐자","리쥬란",
    "스킨보톡스","주름보톡스","턱보톡스","사각턱보톡스","이마보톡스","눈가보톡스",
    "주름","주름개선","주름치료","잔주름","깊은주름","목주름","이마주름","눈가주름",
    "흉터","흉터레이저","여드름흉터","흉터치료","흉터제거","여드름흉터레이저",
    "색소","색소치료","기미","기미레이저","잡티","잡티레이저","홍조","홍조치료",
    "모공","모공레이저","피부결","탄력","탄력레이저","피부리프팅","리프팅레이저",
    "스킨부스터","물광주사","수분주사","엑소좀","줄기세포",
    # 성형외과 시술 (토큰 분해용 + 직접 검색어)
    "코성형","눈성형","쌍꺼풀","쌍꺼풀수술","지방흡입","안면윤곽","지방이식",
    "눈매교정","양악수술","가슴성형","코수술","눈수술","리프팅수술","실리프팅",
    "성형","성형잘하는곳","성형추천","성형외과추천","성형외과잘하는곳",
    # 의료/건강
    "내과","정형외과","소아과","산부인과","안과","이비인후과","신경과","재활의학과",
    "치과","한의원","한방병원","침","추나","도수치료","물리치료",
    "의원","클리닉","병원","요양병원","건강검진",
    # 반려동물
    "동물병원","동물의원","펫샵","펫카페","애완동물","수의사",
    # 강아지
    "애견미용","강아지미용","강아지호텔","애견호텔","강아지위탁","애견위탁","강아지유치원","애견유치원",
    "강아지훈련","애견훈련","강아지산책","강아지돌봄","강아지케어",
    # 고양이
    "고양이호텔","캣호텔","캣스테이","고양이위탁","고양이펜션","고양이돌봄","고양이케어",
    "고양이미용","고양이병원","반려묘호텔","고양이유치원","24시고양이호텔",
    # 공통 펫
    "펫호텔","펫시터","펫케어","펫위탁","반려동물호텔","펫호텔링",
    # 자동차/모터
    "자동차정비","카센터","타이어교체","오일교환","판금도색","자동차검사",
    "세차장","셀프세차","자동세차","디테일링","광택",
    # 생활서비스
    "꽃집","꽃배달","화원","플라워","웨딩","돌잔치","파티",
    "사진관","증명사진","스튜디오","웨딩스튜디오",
    "세탁","빨래방","코인세탁","수선","신발수선",
    "안경","렌즈","안경점","열쇠","자물쇠","인테리어","도배","창호",
    # 숙박/공간대여
    "파티룸","공간대여","모임공간","회의실","연습실","스튜디오대여",
    # 수식어 (전업종 공통)
    "커플데이트","커플","가족","감성","주말","추천","잘하는곳","가성비","후기","리뷰","예약",
    "24시","새벽","당일","무료주차","주차가능","가까운","근처","주변",
]

# 루트 토큰 → 연관 복합 키워드 자동 확장 (keywordList에 없는 조합 생성)
_TOKEN_EXPANSIONS = {
    "과학영재": ["영재학원", "과학학원", "영재교육", "과학교육"],
    "영재":     ["영재학원", "영재교육", "영재수학"],
    "사고력":   ["사고력학원", "사고력수학", "사고력교육"],
    "과학":     ["과학학원", "과학교육"],
    "수학":     ["수학학원"],
    "영어":     ["영어학원"],
    "초등":     ["초등수학학원", "초등학원"],
    "중등":     ["중등수학학원", "중등학원"],
    "유아":     ["유아학원", "유아교육"],
    "입시":     ["입시학원", "입시전문"],
    "내신":     ["내신관리", "내신대비", "내신전문"],
    "논술":     ["논술학원"],
    "헬스":     ["헬스장", "헬스클럽", "피트니스"],
    "PT":       ["개인PT", "퍼스널트레이닝"],
    "필라테스": ["필라테스학원", "필라테스센터"],
    "요가":     ["요가학원", "요가센터"],
    "헤어":     ["미용실", "헤어샵"],
    "피부":     ["피부과", "피부관리", "피부케어"],
    "피부관리": ["피부케어", "피부미용", "에스테틱"],
    "에스테틱": ["에스테틱샵", "피부관리", "피부케어"],
    "여드름":   ["여드름관리", "여드름케어"],
    "윤곽":     ["윤곽관리", "윤곽마사지"],
    "리프팅":   ["리프팅관리", "탄력관리"],
    "웨딩":     ["웨딩관리", "웨딩케어"],
    "고양이":   ["고양이호텔", "캣호텔", "고양이위탁", "고양이케어", "고양이미용"],
    "캣":       ["캣호텔", "캣스테이", "고양이호텔", "고양이위탁"],
    "강아지":   ["강아지호텔", "애견호텔", "강아지위탁", "강아지미용", "강아지돌봄"],
    "애견":     ["애견호텔", "강아지호텔", "애견미용", "애견위탁"],
    "펫":       ["펫호텔", "펫시터", "펫케어"],
    "반려묘":   ["고양이호텔", "캣호텔", "고양이위탁"],
    "반려견":   ["강아지호텔", "애견호텔", "강아지위탁"],
    "계곡":     ["계곡여행", "계곡캠핑", "계곡펜션"],
    "커플":     ["커플캠핑", "커플여행", "커플펜션"],
    "감성":     ["감성캠핑", "감성캠핑장"],
    "카라반":   ["카라반캠핑", "차박"],
    "카페":     ["정원카페", "야경카페", "좋은카페", "대형카페", "베이커리카페", "디저트카페", "카페테라스", "대형베이커리카페"],
    "커피":     ["카페", "브런치카페", "감성카페"],
    "디저트":   ["디저트카페", "베이커리카페"],
    "베이커리": ["베이커리카페", "대형베이커리카페"],
    # 피부과 시술 기기 확장
    "써마지":   ["써마지피부과", "써마지FLX", "피부리프팅"],
    "울쎄라":   ["울쎄라리프팅", "울쎄라피부과"],
    "슈링크":   ["슈링크유니버스", "피부리프팅"],
    "주름":     ["주름보톡스", "주름필러", "주름개선", "주름레이저"],
    "흉터":     ["흉터레이저", "여드름흉터", "흉터치료"],
    "기미":     ["기미레이저", "색소레이저"],
    "모공":     ["모공레이저", "모공치료"],
    # 성형외과
    "성형외과": ["코성형", "눈성형", "쌍꺼풀", "지방흡입", "안면윤곽", "성형외과추천", "성형외과잘하는곳"],
    "성형":     ["성형외과", "코성형", "눈성형", "안면윤곽"],
    "코성형":   ["코수술", "코성형잘하는곳"],
    "눈성형":   ["쌍꺼풀", "눈매교정", "눈성형잘하는곳"],
    "피부과":   ["레이저", "보톡스", "필러", "피부관리"],
}

def _find_tokens_in_kw(kw, locations):
    """keywordList 항목에서 지역 제거 후 의도 토큰 추출 (포함 검색)"""
    remaining = kw.strip()
    # 지역 접두어 제거 (긴 것 우선)
    for loc in sorted(locations, key=len, reverse=True):
        if loc and len(loc) >= 2 and remaining.startswith(loc):
            remaining = remaining[len(loc):]
            break
    if len(remaining) < 2:
        return []
    remaining_lower = remaining.lower()
    found = [t for t in _INTENT_TOKENS if len(t) >= 2 and t.lower() in remaining_lower]
    # 더 긴 토큰의 부분 문자열인 짧은 토큰 제거 (예: "고양이호텔" 있으면 "호텔" 단독 제거)
    found = [t for t in found if not any(t != o and o.endswith(t) for o in found)]
    # 사전에 없으면 버림 (5자 이하 단순어만 fallback 허용 — "아침점심저녁특선" 같은 장문 복합어 차단)
    if not found and len(remaining) <= 5:
        found = [remaining]
    return list(dict.fromkeys(found))

# ─────────────────────────────────────────────────────────────────────────────
def generate_highly_relevant_keywords(store_name, category, address, menu_items, official_keywords, nearby_station="", keyword_list=None, log_func=None):
    locations = []
    clean_name = store_name.strip()
    
    # [v2.9.2의 완벽했던 지점명 파싱 로직 토씨 하나 안 틀리고 복사]
    for suffix in ["본점", "직영점", "지점", "점"]:
        if clean_name.endswith(suffix):
            loc_match = re.search(r'([가-힣a-zA-Z0-9]+)' + suffix + r'$', clean_name)
            if loc_match:
                loc = loc_match.group(1).split()[-1] # 여기서 정확히 '노원', '오산'만 분리!
                locations.extend([loc, f"{loc}역"])
                # 비서울/비경기 주소면 {loc}동도 추가 (서울/경기는 동명 충돌 위험: 예: 노원동=대구 북구)
                if address and not address.startswith("서울") and not address.startswith("경기"):
                    locations.append(f"{loc}동")
                if "호수" in loc: locations.append(loc.replace("호수", ""))
                clean_name = clean_name.replace(loc_match.group(0), "").strip()
            break
            
    # 주소 파싱: 시/구/동에서 위치 추출
    SKIP_CITIES = {"서울", "경기"}  # 너무 광범위한 지역명 제외
    addr_tokens = address.replace(",", " ").split()
    for token in addr_tokens:
        if token.endswith("구") and len(token) > 1:
            gu = token[:-1]                            # 마지막 "구"만 제거 (구로구→구로, replace는 "구"전부제거 버그)
            locations.append(token)                    # "북구" 자체도 추가
            if len(gu) >= 2:
                locations.extend([gu, f"{gu}역"])      # 구로 + 구로역 (1글자 prefix엔 역 추가 안 함: 대역·북역 방지)
        elif token.endswith("군") and len(token) > 1:
            locations.append(token[:-1])               # 홍천군 → 홍천
        elif token.endswith("시") and len(token) > 1:
            si = token[:-1]
            if si not in SKIP_CITIES:
                locations.append(si)                   # 인천시 → 인천
        elif token.endswith("읍") and len(token) > 1:
            locations.append(token[:-1])               # 홍천읍 → 홍천
        elif token.endswith("동") and len(token) > 1 and token not in ["공동", "이동", "감동", "행동"]:
            dong_name = token.replace("동", "")
            locations.append(token)  # 동 이름만 추가 (역 자동생성 안 함 — 허위 역 방지)
            # 산곡3동 → 산곡동도 추가 (행정동 번호 제거)
            base_dong = re.sub(r'\d+$', '', dong_name)
            if base_dong != dong_name and len(base_dong) >= 2:
                locations.append(f"{base_dong}동")
    # 주소 첫 토큰이 시 이름인데 '시' 없이 표기된 경우 (예: "인천 부평구...")
    KNOWN_CITIES = {"인천", "부산", "대구", "대전", "광주", "울산", "세종",
                    "수원", "성남", "안양", "부천", "고양", "용인"}
    if addr_tokens and addr_tokens[0] in KNOWN_CITIES:
        locations.append(addr_tokens[0])
    # 비수도권 도(道) 단위 광역 지역 추가 — 캠핑/여행/숙박 업종에만 (식당·헬스장 등 로컬 업종 제외)
    _PROV_MAP = {"강원": "강원도", "경남": "경상남도", "경북": "경상북도",
                 "전남": "전라남도", "전북": "전라북도", "충남": "충청남도",
                 "충북": "충청북도", "제주": "제주도"}
    _cat_lower = (category + " " + store_name).lower()
    _is_regional_biz = any(c in _cat_lower for c in
        ['캠핑', '야영', '글램핑', '차박', '펜션', '리조트', '숙박', '게스트하우스',
         '민박', '여행', '관광', '레저', '휴양', '낚시', '자연', '농원', '농장'])
    if _is_regional_biz and addr_tokens and addr_tokens[0] in _PROV_MAP:
        prov_short, prov_long = addr_tokens[0], _PROV_MAP[addr_tokens[0]]
        if prov_short not in locations: locations.append(prov_short)
        if prov_long not in locations: locations.append(prov_long)
    # pcmap 교통정보에서 실제 확인된 역만 locations에 추가
    # 긴 복합역명(신분당신논현역)은 끝부분 단순역명(신논현역) 추출
    if nearby_station and len(nearby_station) >= 3:
        if len(nearby_station) <= 6:
            locations.append(nearby_station)
        else:
            m_st = re.search(r'[가-힣]{2,4}역$', nearby_station)
            if m_st and len(m_st.group()) <= 6:
                locations.append(m_st.group())

    locations = list(dict.fromkeys([l for l in locations if l and len(l) >= 2]))
    for skip in ["서울", "경기"]:
        if skip in locations: locations.remove(skip)
    if not locations: locations = [""] 
    
    # ── 공식 키워드 필터 ──────────────────────────────────────────────────────
    _BAD_PATTERNS = re.compile(r'영업중|영업종료|영업시간|\d{3,}|에영업|시에|분에|\d+시\d*분|휴무|정기휴무|임시휴무|특가|이벤트|한정|첫방문|할인쿠폰|프로모션|레귤러|수퍼스페셜|디럭스|스탠다드|스위트룸|싱글룸|더블룸|트윈룸|\d+만원(?!대)|지인소개|기간증정|\d+회(?:추가|증정)|뭉칠수록|혜택최대')
    _bone_kw_pat = re.compile(r'뼈국밥|뼈해장국|뼈다귀|돼지뼈')
    clean_official = [tag for tag in official_keywords
                      if not _BAD_PATTERNS.search(tag) and not _bone_kw_pat.search(tag) and len(tag) <= 15]
    kw_list = [k.strip() for k in (keyword_list or []) if k and len(k.strip()) >= 2 and not _BAD_PATTERNS.search(k)]

    # ── keywordList에서 추가 지역 토큰 추출 (키워드 앞부분에서만 — "진주호탄동" 같은 연속 지명 차단) ──
    for kw in kw_list:
        # 방식 1: 역/동/구/대 suffix (예: 구일역, 고척동)
        m = re.match(r'^[가-힣]{2,3}(?:역|동|구|대)', kw)
        added_by_method1 = False
        if m:
            extra = m.group()
            # 이미 locations에 있는 행정구(xxx구)에 "동" suffix가 붙은 파생형이면 스킵 (북구+동=북구동 방지)
            is_derived_dong = (extra.endswith("동") and
                               any(extra == loc + "동" for loc in locations if loc.endswith("구")))
            if extra not in locations and not is_derived_dong:
                locations.append(extra)
                added_by_method1 = True

        if not added_by_method1:
            # 방식 2: suffix 없는 도시명 추출 (예: 부천→부천, 팔공산→팔공산)
            # 긴 prefix 우선(3→2자) — "팔공산"(3)이 "팔공"(2)보다 먼저 추가됨
            for plen in (3, 2):
                if len(kw) > plen + 2:
                    prefix = kw[:plen]
                    rest = kw[plen:]
                    # 행정구 파생 "동" 방지 (북구동변동→prefix=북구동)
                    is_derived2 = (prefix.endswith("동") and
                                   any(prefix == loc + "동" for loc in locations if loc.endswith("구")))
                    # 이미 알려진 지명에 수식어가 붙은 가짜 지명 방지 ("북구"가 있으면 "북구밤"·"북구화" 스킵)
                    is_extension2 = any(prefix.startswith(loc) and loc != prefix
                                        for loc in locations if len(loc) >= 2)
                    # prefix가 _INTENT_TOKEN 자체이거나 그 접두사이면 시술어 → location 아님
                    # 역방향도 체크: prefix가 intent 토큰으로 시작하면 음식/업종어 (돼지등→돼지갈비의 돼지로 시작)
                    _is_intent_prefix = (prefix in _INTENT_TOKENS or
                                         any(t.startswith(prefix) and len(t) > len(prefix)
                                             for t in _INTENT_TOKENS) or
                                         any(prefix.startswith(t) and len(t) >= 2 and t != prefix
                                             for t in _INTENT_TOKENS))
                    if (re.match(r'^[가-힣]+$', prefix) and not is_derived2 and not is_extension2
                            and not _is_intent_prefix
                            and any(t in rest for t in _INTENT_TOKENS if len(t) >= 3)):
                        if prefix not in locations:
                            locations.append(prefix)
                        # suffix 없는 단순 지명 → 역 suffix 자동 추가 (천호→천호역)
                        if not prefix.endswith(('역', '동', '구', '시', '군', '읍', '면', '리', '대', '산', '강')):
                            st_cand = prefix + "역"
                            if st_cand not in locations:
                                locations.append(st_cand)
                        break

        # 방식 3: kw 전체에서 역/동 패턴 overlapping 검색 (동변동·침산동 등 중간 위치명 추출)
        # locations에 이미 있는 것도 cands_all에 포함해야 필터 비교가 정확함
        cands3_all = []
        for loc_m in re.finditer(r'(?=([가-힣]{2,3}(?:동|역)))', kw):
            extra3 = loc_m.group(1)
            is_derived3 = (extra3.endswith("동") and
                           any(extra3 == loc + "동" for loc in locations if loc.endswith("구")))
            if not is_derived3:
                cands3_all.append(extra3)
        # 더 짧은 후보를 내포하는 긴 후보 제거 (이미 있는 "동변동"과 비교해 "산동변동" 제거)
        cands3_ok = [c for c in cands3_all if not any(c != o and o in c for o in cands3_all)]
        for extra3 in cands3_ok:
            if extra3 not in locations:
                locations.append(extra3)

        # 방식 4: kw 왼쪽부터 locations 순차 제거 → 나머지에서 intent 토큰 제거 → 남은 2-4글자 위치명 추출
        # 예: "북구동변동금호강디저트카페" → 북구·동변동 순차 제거 → "금호강디저트카페" → "디저트카페" 제거 → "금호강"
        remaining4 = kw
        while remaining4:
            loc_matched = False
            for loc in sorted(locations, key=len, reverse=True):
                if loc and remaining4.startswith(loc):
                    remaining4 = remaining4[len(loc):]
                    loc_matched = True
                    break
            if not loc_matched:
                break
        for t in sorted(_INTENT_TOKENS, key=len, reverse=True):
            if len(t) >= 3 and t in remaining4:
                remaining4 = remaining4.replace(t, '', 1)
                break
        # 추출 chunk 유효성 필터:
        # ① 실제 지명 suffix로 끝나야 함 (역/동/구/산/강/천/호 등 또는 공원/호수)
        # ② 기존 location의 초과집합이면 안 됨 (산동변동 ⊃ 동변동 → 제외)
        _LOC_SFXS = {'역', '동', '구', '산', '강', '천', '호', '읍', '면', '리'}
        _LOC_SFXS2 = {'공원', '호수', '댐', '계곡'}
        for chunk in re.findall(r'[가-힣]{2,4}', remaining4):
            loc_suffix_ok = (chunk[-1] in _LOC_SFXS or chunk[-2:] in _LOC_SFXS2)
            is_superset = any(loc in chunk and loc != chunk for loc in locations if len(loc) >= 2)
            if (chunk not in locations and chunk not in _INTENT_TOKENS
                    and not any(chunk == loc + "동" for loc in locations if loc.endswith("구"))
                    and loc_suffix_ok and not is_superset):
                locations.append(chunk)

    locations = list(dict.fromkeys([l for l in locations if l and len(l) >= 2]))
    for skip in ["서울", "경기"]:
        if skip in locations: locations.remove(skip)
    if not locations: locations = [""]
    if log_func: log_func(f"    ㄴ 📌 위치 토큰: {', '.join(locations)}")

    # ── 1순위: keywordList 그대로 (업체주 직접 최적화한 완성형 키워드) ─────────
    def _has_location(kw):
        return any(loc and loc in kw for loc in locations if loc)
    def _multi_loc(kw):
        return sum(1 for loc in locations if loc and len(loc) >= 2 and loc in kw) >= 2
    def _has_intent(kw):
        """_INTENT_TOKENS 단어가 하나라도 포함돼 있으면 True (검색 의도 있음)"""
        kw_lower = kw.lower()
        return any(t.lower() in kw_lower for t in _INTENT_TOKENS if len(t) >= 2)
    # 직접 검색 조건:
    #   ① 지역명 포함 OR ② 의도 토큰 포함 → 실제 검색 키워드
    #   지역도 없고 의도 토큰도 없으면 업주 홍보문구(지인소개시기간증정, 뭉칠수록혜택최대... 등)
    #   → 직접 검색 제외, 토큰 추출용으로만 사용 (어떤 업체든 자동 필터)
    kws = list(dict.fromkeys(
        k for k in kw_list
        if (len(k) > 4 or _has_location(k))
        and not _multi_loc(k)
        and (_has_location(k) or _has_intent(k))
    ))

    # keywordList 토큰화 → 지역 × 의도토큰 조합 생성
    all_kw_tokens = []
    for kw in kw_list:
        all_kw_tokens.extend(_find_tokens_in_kw(kw, locations))
    all_kw_tokens = list(dict.fromkeys(t for t in all_kw_tokens if len(t) >= 2))

    # 루트 토큰 확장: "영재" → 영재학원/영재교육 등 자동 추가 (keywordList에 없는 조합 보완)
    # 복합 토큰 내 루트 확장: "정원카페" → 루트 "카페" → 야경카페/좋은카페 등 추가
    _seen_tokens = set(all_kw_tokens)
    for t in list(all_kw_tokens):
        for expanded in _TOKEN_EXPANSIONS.get(t, []):
            if expanded not in _seen_tokens:
                all_kw_tokens.append(expanded)
                _seen_tokens.add(expanded)
        for root in _TOKEN_EXPANSIONS:
            if root in t and root != t:
                for expanded in _TOKEN_EXPANSIONS[root]:
                    if expanded not in _seen_tokens:
                        all_kw_tokens.append(expanded)
                        _seen_tokens.add(expanded)

    kws_set = set(kws)
    for loc in locations:
        for token in all_kw_tokens:
            combined = f"{loc} {token}" if loc and loc not in token else token
            if combined not in kws_set:
                kws.append(combined)
                kws_set.add(combined)

    # ── 2순위: 지역 × official_keywords 조합 ─────────────────────────────────
    # keywordList가 있으면 all_kw_tokens 기반 필터 적용
    # (루프탑카페 등 keywordList와 무관한 네이버 자동 태그 차단)
    _kw_filter_set = [t for t in all_kw_tokens if len(t) >= 3]
    for tag in clean_official:
        clean_tag = re.sub(r'[^가-힣a-zA-Z0-9]', '', tag).strip()
        if len(clean_tag) < 2:
            continue
        if kw_list and _kw_filter_set:
            if not any(t in clean_tag or clean_tag in t for t in _kw_filter_set):
                continue
        for loc in locations:
            if loc and loc in clean_tag:
                kws.append(clean_tag)
            else:
                kws.append(f"{loc} {clean_tag}".strip())

    # 메뉴 아이템도 지역과 조합
    # _has_intent 필수: 의도 토큰 없는 항목은 홍보문구로 간주 → 자동 제거 (어떤 업종이든 적용)
    # 예: "지인소개시기간증정", "뭉칠수록혜택최대15할인" → 서비스 토큰 없음 → 자동 탈락
    _MENU_GRADE_SKIP = {"스페셜", "럭셔리", "프리미엄", "베이직", "스탠다드", "기본형", "일반형", "고급형"}
    for menu in menu_items:
        clean_m = re.sub(r'[^가-힣a-zA-Z0-9]', '', menu)
        if (2 <= len(clean_m) <= 12
                and not _BAD_PATTERNS.search(clean_m)
                and clean_m not in _MENU_GRADE_SKIP
                and _has_intent(clean_m)):
            for loc in locations:
                kws.append(f"{loc} {clean_m}".strip() if loc and loc not in clean_m else clean_m)

    # ── 3순위: keywordList·official_keywords 둘 다 없을 때만 카테고리 폴백 ────
    if not kw_list and not clean_official:
        cat_str = (category + " " + store_name).lower()
        fallback = []
        if any(x in cat_str for x in ['헬스', 'pt', '피트니스', '휘트니스']):
            fallback = ["헬스장", "PT", "개인PT", "피트니스"]
        elif any(x in cat_str for x in ['학원', '교육', '영재', '사고력']):
            fallback = ["학원", "영재학원", "사고력수학", "교육센터"]
        elif any(x in cat_str for x in ['캠핑', '야영', '글램핑']):
            fallback = ["캠핑장", "글램핑", "오토캠핑"]
        elif any(x in cat_str for x in ['펜션', '풀빌라', '숙박', '호텔']):
            fallback = ["펜션", "풀빌라", "숙소"]
        elif any(x in cat_str for x in ['병원', '치과', '한의원', '클리닉']):
            fallback = ["병원", "치과", "한의원"]
        elif any(x in cat_str for x in ['미용', '헤어']):
            fallback = ["미용실", "헤어샵", "머리잘하는곳"]
        elif any(x in cat_str for x in ['카페', '커피', '디저트', '베이커리']):
            fallback = ["카페", "커피", "디저트"]
        elif any(x in cat_str for x in ['고기', '갈비', '국밥', '식당', '음식점']):
            fallback = ["맛집", "고기집", "맛있는집"]
        else:
            # 최후 수단: 카테고리 단어 자체 사용
            fallback = [w.strip() for w in category.split(',') if len(w.strip()) >= 2][:3]
            fallback.append("추천")
        for loc in locations:
            for intent in fallback:
                kws.append(f"{loc} {intent}".strip() if loc and loc not in intent else intent)

    # ── 음식점 계열 → 지역 × "맛집"/"맛집추천" 디폴트 추가 ──────────────────
    # 네이버 플레이스에 직접 등록 불가한 키워드라 별도로 추가
    _FOOD_CAT_SIGNALS = [
        '음식점', '한식', '일식', '중식', '양식', '분식', '카페', '커피',
        '베이커리', '제과', '디저트', '아이스크림', '술집', '주점',
        '이자카야', '호프', '치킨', '피자', '햄버거', '패스트푸드', '뷔페',
        '해산물', '횟집', '수산', '고기', '갈비', '국밥', '칼국수',
        '순대', '찌개', '정식', '한정식', '삼계탕', '보쌈', '족발',
        '생선', '두부', '비빔밥', '떡볶이', '김밥', '초밥', '스시',
        '라멘', '우동', '스테이크', '파스타', '샌드위치', '쌀국수',
        '카레', '타코', '케밥', '편의점', '정육점', '식당',
    ]
    _food_check = (category + " " + store_name).lower()
    if any(sig in _food_check for sig in _FOOD_CAT_SIGNALS):
        for loc in locations:
            for food_kw in ["맛집", "맛집추천"]:
                combined = f"{loc} {food_kw}".strip() if loc else food_kw
                kws.append(combined)  # 중복은 하단 dedup에서 처리, 전체 지역 보장

    # ── 중복 제거 + 정렬 (keywordList 순서 고정, 나머지만 역/동/구 우선 정렬) ─
    _kl_text = ''.join(kw_list)
    def sort_weight(kw):
        if kw in set(kw_list): return 1000  # keywordList 항상 최상위
        w = 0
        if "역" in kw: w += 30
        elif "동" in kw: w += 20
        elif "구" in kw: w += 10
        # keywordList에 직접 등장하는 랜드마크/지역명 보너스 (용오름, 팔공산, 금호강 등 역/동/구 없는 지명)
        for loc in locations:
            if loc and len(loc) >= 3 and loc in kw and loc in _kl_text:
                w += 15
                break
        return w

    seen = set()
    deduped = []
    for k in kws:
        if k not in seen:
            seen.add(k)
            deduped.append(k)

    deduped.sort(key=sort_weight, reverse=True)
    return deduped[:100]

# ----------------- [2. 블로그 딥스캐너 v6.9 - 메인카드 순위 보정 + 본문 진입 검사] -----------------
def clean_blog_url(url):
    # 네이버 검색/리다이렉트/PC 블로그 주소를 모바일 블로그 주소로 정리합니다.
    if not url:
        return ""
    try:
        parsed = urllib.parse.urlparse(url)
        qs = urllib.parse.parse_qs(parsed.query)
        for key in ["url", "u", "target"]:
            if key in qs and qs[key]:
                url = qs[key][0]
                break
    except Exception:
        pass

    try:
        url = urllib.parse.unquote(url)
    except Exception:
        pass

    if "blog.naver.com" in url and "m.blog.naver.com" not in url:
        url = url.replace("https://blog.naver.com", "https://m.blog.naver.com")
        url = url.replace("http://blog.naver.com", "https://m.blog.naver.com")
    return url


def normalize_text(text):
    if not text:
        return ""
    return re.sub(r"\s+", "", str(text)).strip()

_AE_TABLE = str.maketrans('베레에게세제테케페헤네메체셰', '배래애개새재태캐패해내매채섀')
def _sn(text):
    """ㅔ→ㅐ 발음 정규화 (베럴짐=배럴짐 동일 처리)"""
    return normalize_text(text).translate(_AE_TABLE)


def extract_address_tokens(address):
    # 주소에서 매칭 신뢰도에 도움 되는 토큰만 추출합니다.
    if not address:
        return []
    raw_tokens = re.split(r"[\s,()\[\]<>/]+", address)
    tokens = []
    for t in raw_tokens:
        t = t.strip()
        if len(t) < 2:
            continue
        if t in ["서울", "경기", "인천", "부산", "대구", "대전", "광주", "울산", "세종"]:
            continue
        if any(t.endswith(suffix) for suffix in ["구", "동", "로", "길", "시", "읍", "면"]):
            tokens.append(t)
        elif re.search(r"\d", t) and len(t) >= 3:
            tokens.append(t)
    return list(dict.fromkeys(tokens))[:8]


async def collect_blog_results(page, keyword, limit=30, log_func=None):
    encoded_kw = urllib.parse.quote(keyword)
    blog_url = f"https://search.naver.com/search.naver?ssc=tab.blog.all&sm=tab_jum&query={encoded_kw}"
    await page.goto(blog_url, wait_until="domcontentloaded", timeout=15000)
    await page.wait_for_timeout(800)

    for _ in range(7):
        prev_height = await page.evaluate("document.body.scrollHeight")
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(350)
        new_height = await page.evaluate("document.body.scrollHeight")
        if new_height == prev_height:
            break

    results = await page.evaluate('''(limit) => {
        function norm(s) {
            return (s || "").replace(/[\\s]+/g, " ").trim();
        }
        function cleanHref(href) {
            try { return decodeURIComponent(href || ""); } catch(e) { return href || ""; }
        }
        function isMainPostLink(href) {
            href = cleanHref(href);
            if (!href.includes("blog.naver.com")) return false;
            if (href.includes("PostList") || href.includes("profile") || href.includes("section")) return false;
            return /blog\\.naver\\.com\\/([^/?#]+)\\/\\d{9,13}/.test(href) ||
                   /blog\\.naver\\.com\\/PostView/.test(href) ||
                   (href.includes("blogId=") && href.includes("logNo="));
        }
        function getPostKey(href) {
            href = cleanHref(href);
            let m = href.match(/blog\\.naver\\.com\\/([^/?#]+)\\/(\\d{9,13})/);
            if (m) return m[1] + "/" + m[2];
            try {
                const u = new URL(href);
                const blogId = u.searchParams.get("blogId") || "";
                const logNo = u.searchParams.get("logNo") || "";
                if (blogId && logNo) return blogId + "/" + logNo;
            } catch(e) {}
            return href.split("?")[0];
        }
        function isInsideAd(el) {
            let p = el;
            for (let i = 0; i < 8 && p; i++, p = p.parentElement) {
                const cls = ((p.className || "") + " " + (p.id || "")).toString().toLowerCase();
                if (/power_link|sp_nreview|ad_area|_ad|ugcad|adchoice/.test(cls)) return true;
            }
            return false;
        }
        function isTitleLike(txt) {
            if (!txt || txt.length < 8 || txt.length > 150) return false;
            if (txt.includes("blog.naver") || txt.includes("naver.com") || txt.includes("›")) return false;
            if (/^https?:/.test(txt)) return false;
            return true;
        }
        function extractTitle(card) {
            // 1순위: 제목 전용 셀렉터 (URL 포함 텍스트 제외)
            const selectors = [
                "strong.total_tit", "a.api_txt_lines.total_tit",
                ".total_tit", "a.title_link", "a.link_tit",
                ".title_area a", "strong a", "h2 a", ".article_tit"
            ];
            for (const sel of selectors) {
                for (const el of Array.from(card.querySelectorAll(sel))) {
                    const txt = norm(el.innerText || el.textContent || "");
                    if (isTitleLike(txt)) return txt;
                }
            }
            // 2순위: 포스트 링크 앵커에서 (URL·닉네임 패턴 제외)
            for (const a of Array.from(card.querySelectorAll("a"))) {
                const href = a.href || a.getAttribute("href") || "";
                if (!isMainPostLink(href)) continue;
                const txt = norm(a.innerText || a.textContent || "").replace(/blog[.]naver[.]com[^ ]*/g, "").trim();
                if (isTitleLike(txt)) return txt;
            }
            // 3순위: 카드 전체 텍스트에서 제일 긴 줄 추출
            const lines = (card.innerText || "").split(/[\\n\\r]+/).map(l => l.trim()).filter(l => isTitleLike(l));
            lines.sort((a, b) => b.length - a.length);
            return lines[0] ? lines[0].slice(0, 100) : "";
        }

        const seenPosts = new Set();
        const list = [];

        function collectFromCard(card) {
            if (isInsideAd(card)) return;
            const anchors = Array.from(card.querySelectorAll("a[href]"));
            const postLinks = anchors
                .map(a => a.getAttribute("href") || "")
                .filter(isMainPostLink);
            if (postLinks.length === 0) return;
            const key = getPostKey(postLinks[0]);
            if (seenPosts.has(key)) return;
            seenPosts.add(key);
            const allLinks = anchors.map(a => a.href || a.getAttribute("href") || "");
            list.push({
                rank: list.length + 1,
                title: extractTitle(card),
                link: postLinks[0],
                text: norm(card.innerText || ""),
                cardLinks: allLinks,
                card_html: card.outerHTML.slice(0, 8000)
            });
        }

        // 1순위: ugcItem (개별 포스팅 카드, DOM 순서 = 실제 순위) — Naver SDS 신 디자인
        for (const card of document.querySelectorAll('div[data-template-id="ugcItem"]')) {
            if (list.length >= limit) break;
            collectFromCard(card);
        }

        // 2순위: fds-web-doc-root (상위 피처드 카드, ugcItem 없을 때 폴백)
        if (list.length < 5) {
            for (const card of document.querySelectorAll('div[class*="fds-web-doc-root"]')) {
                if (list.length >= limit) break;
                collectFromCard(card);
            }
        }

        // 3순위: data-template-id="layout" 카드 (폴백)
        if (list.length < 5) {
            for (const card of document.querySelectorAll('div[data-template-id="layout"]')) {
                if (list.length >= limit) break;
                collectFromCard(card);
            }
        }

        // 4순위: 구 디자인 li.bx / .view_wrap / .total_wrap 폴백
        if (list.length < 5) {
            for (const card of document.querySelectorAll("li.bx, .view_wrap, .total_wrap")) {
                if (list.length >= limit) break;
                collectFromCard(card);
            }
        }

        // 4순위: 전체 앵커 스캔 (최후 수단)
        if (list.length < 5) {
            const allAnchors = Array.from(document.querySelectorAll("a[href*='blog.naver.com']"));
            for (const a of allAnchors) {
                if (list.length >= limit) break;
                const href = a.getAttribute("href") || "";
                if (!isMainPostLink(href)) continue;
                if (isInsideAd(a)) continue;
                const key = getPostKey(href);
                if (seenPosts.has(key)) continue;
                seenPosts.add(key);
                // 가장 가까운 블록 조상에서 제목 추출 시도
                let card = a.parentElement;
                for (let i = 0; i < 6 && card && card !== document.body; i++) {
                    const hasSibs = card.parentElement &&
                        Array.from(card.parentElement.children).filter(c => c !== card).length > 0;
                    if (hasSibs) break;
                    card = card.parentElement;
                }
                const titleTxt = card ? norm(card.innerText || "").split(/[\\n\\r]+/)[0].slice(0, 80) : "";
                list.push({
                    rank: list.length + 1,
                    title: (isTitleLike(titleTxt) ? titleTxt : null) ||
                           norm(a.innerText || a.textContent || "").slice(0, 80) || "제목 확인 필요",
                    link: href,
                    text: card ? norm(card.innerText || "").slice(0, 200) : "",
                    cardLinks: [href],
                    card_html: card ? card.outerHTML.slice(0, 4000) : ""
                });
            }
        }

        return list;
    }''', limit)

    return results

async def inspect_blog_post(page, blog_url, store_name, place_id, address, card_text="", card_links=None, card_html=""):
    card_links = card_links or []
    pid = str(place_id) if place_id else ""
    clean_store = normalize_text(store_name)
    address_tokens = extract_address_tokens(address)

    brand_candidates = []
    for part in re.split(r"\s+", store_name.strip()):
        original = part
        part = re.sub(r"(본점|직영점|지점|점)$", "", part).strip()
        suffix_removed = (original != part)
        if len(part) < 3:
            continue
        # 지점 suffix 제거 후 4자 이하 순한글이면 지역명일 가능성 높음 → 브랜드 후보 제외
        # 예: "광화문점"→"광화문"(3자), "강남점"→"강남"(2자) 모두 지역명
        if suffix_removed and len(part) <= 4 and re.match(r'^[가-힣]+$', part):
            continue
        brand_candidates.append(part)
    if store_name.strip():
        brand_candidates.append(store_name.strip())
    brand_candidates = list(dict.fromkeys(brand_candidates))

    score = 0
    reasons = []

    # 1차: 카드 링크 및 card_html에서 place_id 즉시 감지
    joined_card_links = " ".join(card_links) + " " + card_html
    if pid and pid in joined_card_links:
        score += 120
        reasons.append("검색카드 장소태그ID")

    if clean_store and (clean_store in normalize_text(card_text) or _sn(clean_store) in _sn(card_text)):
        score += 35
        reasons.append("검색카드 업체명")

    # 카드에서 이미 합격 점수면 본문 진입 없이 조기 반환
    if score >= 100:
        return {"matched": True, "score": score, "reasons": reasons, "text_sample": "", "page_title": ""}

    try:
        url = clean_blog_url(blog_url)
        await page.goto(url, wait_until="domcontentloaded", timeout=12000)
        await page.wait_for_timeout(350)

        # 실제 페이지 제목 추출 (카드 제목 오류 보정용)
        page_title = ""
        try:
            raw_title = await page.evaluate("() => document.title")
            for sep in [" : ", " | ", " - ", "::","｜"]:
                if sep in raw_title:
                    raw_title = raw_title.split(sep)[0].strip()
                    break
            page_title = raw_title[:100].strip()
        except Exception:
            pass

        for _ in range(2):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(200)

        # outer HTML + iframe 전수 검색
        outer_html = await page.evaluate("() => document.documentElement.outerHTML")
        html_decoded = urllib.parse.unquote(outer_html)

        # iframe 내부도 검색
        iframe_html = ""
        try:
            for frame in page.frames:
                if frame == page.main_frame:
                    continue
                try:
                    fhtml = await frame.evaluate("() => document.documentElement.outerHTML")
                    iframe_html += fhtml
                except Exception:
                    pass
        except Exception:
            pass

        full_html = html_decoded + " " + iframe_html

        if pid:
            pid_patterns = [
                pid,
                f"placeId/{pid}",
                f'placeId":"{pid}',
                f"mapId={pid}",
                f"pid={pid}",
                f"place%2F{pid}",
                urllib.parse.quote(f'"placeId":"{pid}"'),
            ]
            for pat in pid_patterns:
                if pat in full_html:
                    score += 150
                    reasons.append("본문 플레이스ID")
                    break

        map_signals = [
            "map.naver.com", "place.naver.com", "pcmap.place.naver.com", "naver.me",
            "placeSection", "placeId", "businessId", "blog_map", "se-map"
        ]
        if any(sig in full_html for sig in map_signals):
            score += 25
            reasons.append("본문 지도/장소태그 흔적")

        body_text = await page.evaluate("() => (document.body ? document.body.innerText : '').replace(/[ \\t\\n\\r]+/g, ' ').trim()")
        normalized_body = normalize_text(body_text)

        _sn_body = _sn(body_text)
        if clean_store and (_sn(clean_store) in _sn_body):
            score += 55
            reasons.append("본문 업체명")
        else:
            matched_brands = [b for b in brand_candidates if normalize_text(b) and _sn(b) in _sn_body]
            if matched_brands:
                score += 30
                reasons.append(f"본문 브랜드명({matched_brands[0]})")

        matched_addr = [t for t in address_tokens if t and t in body_text]
        if len(matched_addr) >= 2:
            score += 35
            reasons.append("본문 주소일부")
        elif len(matched_addr) == 1:
            score += 15
            reasons.append("본문 주소1개")

        has_pid_in_body = "본문 플레이스ID" in reasons
        has_store_in_body = "본문 업체명" in reasons or any(r.startswith("본문 브랜드명") for r in reasons)
        has_strong_signal = has_pid_in_body or "본문 업체명" in reasons or "검색카드 업체명" in reasons
        has_brand_signal = any(r.startswith("본문 브랜드명") for r in reasons)

        # ── 제목 최종 관문 ─────────────────────────────────────────────────────
        # pid 확인 기준:
        #   강함 — 카드 장소태그ID or 카드 업체명 or (본문 pid + 본문 업체명)
        #   약함 — 본문 pid만 (부수적장소태그 가능성)
        # 고유 브랜드가 있을 때만 경쟁사 오탐 차단 로직 적용
        _card_ok = "검색카드 장소태그ID" in reasons or "검색카드 업체명" in reasons
        _unique_brands = [
            b for b in brand_candidates
            if b != store_name.strip()
            and b not in _INTENT_TOKENS
            and len(normalize_text(b)) >= 3
            and not (re.match(r'^\d+', b) and len(normalize_text(b)) <= 3)
        ]
        # 강한 확인: 카드 or 본문 pid (부수적장소태그는 아래 step 2에서 별도 차단)
        _strong_confirm = _card_ok or has_pid_in_body

        _sn_title = _sn(page_title) if page_title else ""
        if not _strong_confirm:
            # pid 미확인 → 고유 브랜드 있으면 제목에 브랜드 필수 (경쟁사 오탐 차단)
            if _unique_brands and page_title:
                brand_in_title = any(
                    _sn(b) in _sn_title
                    for b in _unique_brands if len(normalize_text(b)) >= 2
                )
                if not brand_in_title:
                    reasons.append("제목업체명없음(부수적언급)")
                    return {"matched": False, "score": score, "reasons": reasons,
                            "text_sample": "", "page_title": page_title}
            elif not _unique_brands:
                # 업종명=상호명(고양이호텔 등): pid 없고 카드업체명도 없을 때만 제목 체크
                if not has_pid_in_body and "검색카드 업체명" not in reasons and page_title:
                    brand_in_title = _sn(store_name.strip()) in _sn_title
                    if not brand_in_title:
                        reasons.append("제목업체명없음(부수적언급)")
                        return {"matched": False, "score": score, "reasons": reasons,
                                "text_sample": "", "page_title": page_title}

        # 본문 pid 있지만 본문 업체명 없고 카드 미검출 → 부수적장소태그 여부 확인
        if has_pid_in_body and not has_store_in_body and not _card_ok:
            brand_in_title = page_title and any(_sn(b) in _sn_title for b in brand_candidates)
            if not brand_in_title:
                reasons.append("부수적장소태그(제목업체명없음)")
                return {"matched": False, "score": score, "reasons": reasons, "text_sample": body_text[:120], "page_title": page_title}

        if has_strong_signal:
            threshold = 55
        elif has_brand_signal:
            threshold = 75
        else:
            threshold = 130
        matched = score >= threshold
        return {
            "matched": matched,
            "score": score,
            "reasons": reasons,
            "text_sample": body_text[:120],
            "page_title": page_title
        }

    except Exception as e:
        matched = score >= 100
        if matched:
            reasons.append("본문진입실패/카드근거충분")
        else:
            reasons.append(f"본문검사오류:{str(e)[:40]}")
        return {
            "matched": matched,
            "score": score,
            "reasons": reasons,
            "text_sample": "",
            "page_title": ""
        }


async def check_blog_ranking_deep(page, inspect_page, keyword, store_name, place_id, address, log_func=None, max_hits=5, url_cache=None):
    def _log(msg):
        if log_func:
            log_func(msg)

    pid_str = str(place_id) if place_id else ""
    if url_cache is None:
        url_cache = {}
    try:
        results = await collect_blog_results(page, keyword, limit=10, log_func=log_func)
        if not results:
            return [{"status": "검색결과없음", "rank": None, "title": "", "blog_link": "", "score": 0, "reasons": []}]

        hits = []
        for item in results:
            if len(hits) >= max_hits:
                break

            # 1차: card_html + cardLinks에서 place_id 즉시 감지
            card_combined = item.get("card_html", "") + " ".join(item.get("cardLinks", []))
            if pid_str and pid_str in card_combined:
                _log(f"      ↳ {item['rank']}위 카드즉시감지 → 합격")
                hits.append({
                    "status": f"{item['rank']}위",
                    "rank": item["rank"],
                    "title": item.get("title", ""),
                    "blog_link": clean_blog_url(item.get("link", "")),
                    "score": 150,
                    "reasons": ["카드장소태그ID"]
                })
                continue

            # 2차: URL 캐시 확인 → 캐시 히트시 재검사 생략
            item_url = clean_blog_url(item.get("link", ""))
            if item_url and item_url in url_cache:
                checked = url_cache[item_url]
                _log(f"      ↳ {item['rank']}위 캐시({item.get('title','')[:15]}) 점수={checked['score']}")
            else:
                # 3차: 본문 딥스캔
                _log(f"      ↳ {item['rank']}위 딥스캔 중... ({item.get('title','')[:20]})")
                checked = await inspect_blog_post(
                    page=inspect_page,
                    blog_url=item.get("link", ""),
                    store_name=store_name,
                    place_id=place_id,
                    address=address,
                    card_text=item.get("text", ""),
                    card_links=item.get("cardLinks", []),
                    card_html=item.get("card_html", "")
                )
                if item_url:
                    url_cache[item_url] = checked
                _log(f"         점수={checked['score']} 근거={checked['reasons']}")

            if checked["matched"]:
                _log(f"      ↳ {item['rank']}위 → 합격")
                # page_title이 있으면 카드 제목보다 우선 사용 (닉네임+URL 오류 보정)
                best_title = checked.get("page_title") or item.get("title", "")
                hits.append({
                    "status": f"{item['rank']}위",
                    "rank": item["rank"],
                    "title": best_title,
                    "blog_link": item_url,
                    "score": checked["score"],
                    "reasons": checked["reasons"]
                })

        if not hits:
            return [{"status": "순위권 밖", "rank": None, "title": "", "blog_link": "", "score": 0, "reasons": []}]

        hits.sort(key=lambda x: x["rank"] if x["rank"] else 999)
        return hits

    except Exception as e:
        return [{"status": "추적오류", "rank": None, "title": "", "blog_link": "", "score": 0, "reasons": [str(e)[:80]]}]

# ----------------- [UI 대시보드 및 메인 로직] -----------------
class PremiumMarketingApp:
    def __init__(self, root, auth=None):
        self.root = root
        self.auth = auth
        self.root.title(f"플레이스 마스터 PRO (v{APP_VERSION})")
        self.root.geometry("1100x740")
        self.root.resizable(True, True)
        self.root.minsize(700, 500)
        self.root.configure(bg="#0b1329")

        self.config = load_config()
        self.is_running = False
        self.stop_requested = False
        self.analysis_mode = "place"
        self.eta_string = ""

        # 로그인된 경우 Firestore에서 매장 목록 + 분석기록 동기화
        if auth:
            try:
                remote_stores = auth.load_user_stores()
                if remote_stores is not None:
                    self.config['stores'] = remote_stores
                else:
                    self.config['stores'] = []
                    auth.save_user_stores([])
                save_config(self.config)
            except Exception:
                pass
            try:
                remote_hist = auth.load_user_history()
                local_hist  = load_history()
                # 로컬에 있고 Firebase에 없는 업체 → 백그라운드로 업로드 (기존 기록 보호)
                _missing = [sn for sn in local_hist if sn not in remote_hist]
                if _missing:
                    def _upload_local_hist(a, hist, names):
                        import threading
                        def _run():
                            for sn in names:
                                try:
                                    a.save_user_history_store(sn, hist[sn])
                                except Exception:
                                    pass
                        threading.Thread(target=_run, daemon=True).start()
                    _upload_local_hist(auth, local_hist, _missing)
                # Firebase에 있고 로컬에 없는 날짜 → 병합
                if remote_hist:
                    merged = merge_history(local_hist, remote_hist)
                    if merged != local_hist:
                        save_history(merged)
            except Exception:
                pass

        self.setup_styles()
        self.setup_ui()
        self.refresh_store_table()
        
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#1c2541", foreground="#f8fafc", fieldbackground="#1c2541", rowheight=28, font=("Malgun Gothic", 9))
        style.map("Treeview", background=[("selected", "#3a86ff")], foreground=[("selected", "#ffffff")])
        style.configure("Treeview.Heading", background="#0b1329", foreground="#f8fafc", font=("Malgun Gothic", 9, "bold"))
        style.configure("TCombobox", fieldbackground="#243256", background="#243256", foreground="#f8fafc")
        style.map("TCombobox", fieldbackground=[("readonly", "#243256")], foreground=[("readonly", "#f8fafc")])
        style.configure("Hist.TCombobox", fieldbackground="#334155", background="#334155", foreground="#f8fafc", selectbackground="#3a86ff", selectforeground="#ffffff", arrowcolor="#f8fafc")
        style.map("Hist.TCombobox", fieldbackground=[("readonly", "#334155")], background=[("readonly", "#334155")], foreground=[("readonly", "#f8fafc")], arrowcolor=[("readonly", "#f8fafc")])
        style.configure("TCheckbutton", background="#0b1329", foreground="#cbd5e1", font=("Malgun Gothic", 9))
        style.configure("TNotebook", background="#0b1329", borderwidth=0)
        # 비활성 탭: #2a3f6a 배경 + 밝은 회색 글자 (비활성 느낌, 항상 잘 보임)
        # selected/active는 map으로 override, 나머지(비활성)는 configure 기본값 사용
        style.configure("TNotebook.Tab",
                        background="#2a3f6a", foreground="#a8bede",
                        padding=[7, 5], font=("Malgun Gothic", 9))
        style.map("TNotebook.Tab",
                  background=[("selected", "#3a86ff"), ("active", "#3a5fa0")],
                  foreground=[("selected", "white"), ("active", "#e2e8f0")])
        
    def setup_ui(self):
        # ── 스크롤 가능 캔버스 래퍼 ──────────────────────────────────────
        _h_bar = ttk.Scrollbar(self.root, orient="horizontal")
        _v_bar = ttk.Scrollbar(self.root, orient="vertical")
        _h_bar.pack(side="bottom", fill="x")
        _v_bar.pack(side="right", fill="y")
        _canvas = tk.Canvas(self.root, bg="#0b1329",
                            xscrollcommand=_h_bar.set, yscrollcommand=_v_bar.set,
                            highlightthickness=0)
        _canvas.pack(side="left", fill="both", expand=True)
        _h_bar.config(command=_canvas.xview)
        _v_bar.config(command=_canvas.yview)
        inner = tk.Frame(_canvas, bg="#0b1329")
        _cwin = _canvas.create_window((0, 0), window=inner, anchor="nw")

        def _fit_canvas(e):
            req_w = inner.winfo_reqwidth()
            req_h = inner.winfo_reqheight()
            _canvas.itemconfigure(_cwin,
                                  width=max(e.width, req_w),
                                  height=max(e.height, req_h))
            _canvas.configure(scrollregion=_canvas.bbox("all"))
        _canvas.bind("<Configure>", _fit_canvas)
        inner.bind("<Configure>", lambda e: _canvas.configure(scrollregion=_canvas.bbox("all")))

        # ── 기존 UI — 부모: self.root → inner ───────────────────────────
        header_frame = tk.Frame(inner, bg="#1c2541", height=70)
        header_frame.pack(fill="x", side="top")
        tk.Label(header_frame, text="플레이스 마스터 PRO", font=("Malgun Gothic", 16, "bold"), fg="#3a86ff", bg="#1c2541").pack(side="left", padx=20, pady=15)
        tk.Label(header_frame, text=f"v{APP_VERSION} · 네이버 플레이스 & 블로그 실시간 순위 추적 시스템", font=("Malgun Gothic", 9), fg="#64748b", bg="#1c2541").pack(side="left", pady=22, padx=(5, 0))

        container = tk.Frame(inner, bg="#0b1329")
        container.pack(fill="both", expand=True, padx=15, pady=15)

        left_pane = tk.Frame(container, bg="#0b1329")
        left_pane.pack(side="left", fill="y", anchor="nw", padx=(0, 20))
        right_pane = tk.Frame(container, bg="#0b1329")
        right_pane.pack(side="left", fill="both", expand=True)
        
        reg_box = tk.LabelFrame(left_pane, text=" 📍 신규 광고주 계정 등록 ", font=("Malgun Gothic", 10, "bold"), bg="#1c2541", fg="#f8fafc", bd=1, padx=15, pady=15)
        reg_box.pack(fill="x", pady=(0, 15))
        
        tk.Label(reg_box, text="광고주 업체명", font=("Malgun Gothic", 9, "bold"), fg="#cbd5e1", bg="#1c2541", width=12, anchor="w").grid(row=0, column=0, sticky="w", pady=10)
        self.ent_store_name = tk.Entry(reg_box, width=38, bg="#090d1a", fg="#f8fafc", insertbackground="#3a86ff", bd=1, relief="solid", font=("Malgun Gothic", 10))
        self.ent_store_name.grid(row=0, column=1, pady=10, padx=(10, 0), sticky="ew")
        
        tk.Label(reg_box, text="플레이스 URL", font=("Malgun Gothic", 9, "bold"), fg="#cbd5e1", bg="#1c2541", width=12, anchor="w").grid(row=1, column=0, sticky="w", pady=10)
        self.ent_place_url = tk.Entry(reg_box, width=38, bg="#090d1a", fg="#f8fafc", insertbackground="#3a86ff", bd=1, relief="solid", font=("Malgun Gothic", 10))
        self.ent_place_url.grid(row=1, column=1, pady=10, padx=(10, 0), sticky="ew")

        tk.Label(reg_box, text="키워드추가", font=("Malgun Gothic", 9, "bold"), fg="#cbd5e1", bg="#1c2541", width=12, anchor="nw").grid(row=2, column=0, sticky="nw", pady=(8, 0))
        _kw_wrap = tk.Frame(reg_box, bg="#1c2541")
        _kw_wrap.grid(row=2, column=1, pady=(8, 0), padx=(10, 0), sticky="ew")
        self.txt_custom_kw = tk.Text(_kw_wrap, width=36, height=3, bg="#090d1a", fg="#f8fafc", insertbackground="#3a86ff", bd=1, relief="solid", font=("Malgun Gothic", 9), wrap="word")
        self.txt_custom_kw.pack(side="left", fill="both", expand=True)
        _kw_scroll = ttk.Scrollbar(_kw_wrap, orient="vertical", command=self.txt_custom_kw.yview)
        _kw_scroll.pack(side="right", fill="y")
        self.txt_custom_kw.config(yscrollcommand=_kw_scroll.set)
        _kw_hint = tk.Frame(reg_box, bg="#1c2541")
        _kw_hint.grid(row=3, column=1, sticky="ew", padx=(10, 0), pady=(2, 0))
        tk.Label(_kw_hint, text="쉼표·줄바꿈 구분 | 30위 내 노출 시 순위 표기", font=("Malgun Gothic", 7), fg="#64748b", bg="#1c2541").pack(side="left")
        tk.Button(_kw_hint, text="💾 키워드 저장", font=("Malgun Gothic", 8, "bold"), bg="#334155", fg="#f8fafc", bd=0, padx=8, pady=2, cursor="hand2", command=self.save_custom_keywords).pack(side="right")

        ai_card = tk.Frame(reg_box, bg="#090d1a", bd=1, relief="solid", padx=10, pady=12)
        ai_card.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(15, 10))
        tk.Label(ai_card, text="🔍 스마트 키워드 자동 생성 · 블로그 장소태그 딥스캔", font=("Malgun Gothic", 9, "bold"), fg="#3a86ff", bg="#090d1a").pack(anchor="w")
        tk.Label(ai_card, text="업종·지역·메뉴를 분석해 최적 키워드를 자동 생성합니다.\n블로그 포스팅의 장소 ID를 정밀 검출하여 정확한 순위를 제공합니다.", font=("Malgun Gothic", 8), fg="#94a3b8", bg="#090d1a", justify="left").pack(anchor="w", pady=(5, 0))

        btn_grp = tk.Frame(reg_box, bg="#1c2541")
        btn_grp.grid(row=5, column=0, columnspan=2, pady=(10, 0), sticky="ew")
        self.btn_add = tk.Button(btn_grp, text="➕ 신규 등록", font=("Malgun Gothic", 9, "bold"), bg="#3a86ff", fg="white", bd=0, width=22, height=2, cursor="hand2", command=self.add_store)
        self.btn_add.pack(side="left", padx=(5, 5), fill="x", expand=True)
        self.btn_delete = tk.Button(btn_grp, text="❌ 선택 계정 삭제", font=("Malgun Gothic", 9, "bold"), bg="#f43f5e", fg="white", bd=0, width=22, height=2, cursor="hand2", command=self.delete_store)
        self.btn_delete.pack(side="right", padx=(5, 5), fill="x", expand=True)
        
        db_box = tk.LabelFrame(left_pane, text=" 👥 관리 대상 광고주 목록 ", font=("Malgun Gothic", 10, "bold"), bg="#1c2541", fg="#f8fafc", bd=1, padx=10, pady=10)
        db_box.pack(fill="x")
        self.tree = ttk.Treeview(db_box, columns=("name", "url"), show="headings", height=8)
        self.tree.heading("name", text="광고주명"); self.tree.heading("url", text="플레이스 URL 주소")
        self.tree.column("name", width=120, anchor="center"); self.tree.column("url", width=340, anchor="w")
        scrollbar = ttk.Scrollbar(db_box, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True); scrollbar.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewSelect>>", self.on_treeview_select)
        
        ctrl_box = tk.LabelFrame(right_pane, text=" ⚙️ 마케팅 자동화 옵션 설정 ", font=("Malgun Gothic", 10, "bold"), bg="#1c2541", fg="#f8fafc", bd=1, padx=15, pady=10)
        ctrl_box.pack(fill="x", pady=(0, 15))
        opt_grid = tk.Frame(ctrl_box, bg="#1c2541")
        opt_grid.pack(fill="x")
        
        self.var_safe_mode = tk.BooleanVar(value=self.config.get("safe_mode", True))
        ttk.Checkbutton(opt_grid, text="🛡️ 스텔스 안전 우회 모드", variable=self.var_safe_mode).grid(row=0, column=0, sticky="w", pady=3)
        tk.Label(opt_grid, text="카톡 보고서 스타일:", font=("Malgun Gothic", 9, "bold"), fg="#cbd5e1", bg="#1c2541").grid(row=0, column=1, sticky="e", padx=(20, 5))
        self.cmb_kakao_style = ttk.Combobox(opt_grid, values=["이모지 Trendy형", "비즈니스 격식형", "초간단 핵심요약형"], width=15, state="readonly")
        self.cmb_kakao_style.set(self.config.get("kakao_style", "이모지 Trendy형"))
        self.cmb_kakao_style.grid(row=0, column=2, sticky="w")
        
        run_control_frame = tk.Frame(right_pane, bg="#0b1329")
        run_control_frame.pack(fill="x", pady=(0, 15))
        frame_run_buttons = tk.Frame(run_control_frame, bg="#0b1329")
        frame_run_buttons.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        self.btn_place_run = tk.Button(frame_run_buttons, text="📍 플레이스 분석", font=("Malgun Gothic", 11, "bold"), bg="#10b981", fg="white", bd=0, height=2, cursor="hand2", command=self.start_analysis_place)
        self.btn_place_run.pack(side="left", fill="x", expand=True, padx=2.5)
        self.btn_blog_run = tk.Button(frame_run_buttons, text="✍️ 블로그 분석", font=("Malgun Gothic", 11, "bold"), bg="#10b981", fg="white", bd=0, height=2, cursor="hand2", command=self.start_analysis_blog)
        self.btn_blog_run.pack(side="left", fill="x", expand=True, padx=2.5)
        
        self.btn_stop = tk.Button(run_control_frame, text="🛑 분석 중지", font=("Malgun Gothic", 11, "bold"), bg="#64748b", fg="white", bd=0, height=2, state="disabled", cursor="hand2", command=self.request_stop)
        self.btn_stop.pack(side="right", fill="x", expand=True, padx=(5, 0))

        bulk_bar = tk.Frame(right_pane, bg="#162040", pady=5)
        bulk_bar.pack(fill="x", pady=(0, 8))
        self.var_bulk_place = tk.BooleanVar(value=False)
        self.var_bulk_blog = tk.BooleanVar(value=False)
        tk.Checkbutton(bulk_bar, text="  📋 플레이스 전체 광고주 일괄분석", variable=self.var_bulk_place,
                       bg="#162040", fg="#cbd5e1", selectcolor="#0b1329",
                       activebackground="#162040", activeforeground="#ffffff",
                       font=("Malgun Gothic", 9), cursor="hand2").pack(side="left", padx=(10, 20))
        tk.Checkbutton(bulk_bar, text="  📋 블로그 전체 광고주 일괄분석", variable=self.var_bulk_blog,
                       bg="#162040", fg="#cbd5e1", selectcolor="#0b1329",
                       activebackground="#162040", activeforeground="#ffffff",
                       font=("Malgun Gothic", 9), cursor="hand2").pack(side="left", padx=0)

        tab_control = ttk.Notebook(right_pane)
        self.tab_log = tk.Frame(tab_control, bg="#0f172a")
        self.status_bar = tk.Frame(self.tab_log, bg="#1e293b", height=35)
        self.status_bar.pack(fill="x", side="top")
        self.lbl_status_led = tk.Label(self.status_bar, text="●", font=("Malgun Gothic", 12, "bold"), fg="#ef4444", bg="#1e293b")
        self.lbl_status_led.pack(side="left", padx=(15, 5), pady=5)
        self.lbl_status_text = tk.Label(self.status_bar, text="엔진 대기 중 (정지됨)", font=("Malgun Gothic", 9, "bold"), fg="#94a3b8", bg="#1e293b")
        self.lbl_status_text.pack(side="left", padx=5, pady=5)
        btn_copy_log = tk.Button(self.status_bar, text=" 📋 로그 복사 ", font=("Malgun Gothic", 8, "bold"), bg="#334155", fg="#94a3b8", bd=0, padx=8, pady=3, cursor="hand2", relief="flat", command=self.copy_log_to_clipboard)
        btn_copy_log.pack(side="right", padx=10, pady=5)
        self.txt_log = scrolledtext.ScrolledText(self.tab_log, bg="#0f172a", fg="#38bdf8", font=("Consolas", 10), bd=0, insertbackground="#38bdf8")
        self.txt_log.pack(fill="both", expand=True)
        # 우클릭 컨텍스트 메뉴
        log_menu = tk.Menu(self.txt_log, tearoff=0)
        log_menu.add_command(label="복사", command=lambda: self.txt_log.event_generate("<<Copy>>"))
        log_menu.add_command(label="전체 선택", command=lambda: (self.txt_log.tag_add("sel", "1.0", tk.END), self.txt_log.mark_set(tk.INSERT, "1.0"), self.txt_log.see(tk.INSERT)))
        log_menu.add_separator()
        log_menu.add_command(label="전체 복사", command=self.copy_log_to_clipboard)
        self.txt_log.bind("<Button-3>", lambda e: log_menu.tk_popup(e.x_root, e.y_root))

        self.tab_kakao = tk.Frame(tab_control, bg="#f8fafc")
        copy_bar = tk.Frame(self.tab_kakao, bg="#1c2541", height=38)
        copy_bar.pack(fill="x", side="top")
        tk.Label(copy_bar, text="📲  보고서 완성 · 카톡으로 전송하세요", font=("Malgun Gothic", 9, "bold"), fg="#94a3b8", bg="#1c2541").pack(side="left", padx=12, pady=8)
        self.btn_copy = tk.Button(copy_bar, text="  📋 전체 복사  ", font=("Malgun Gothic", 9, "bold"), bg="#3a86ff", fg="#ffffff", bd=0, padx=12, pady=4, cursor="hand2", relief="flat", command=self.copy_to_clipboard)
        self.btn_copy.pack(side="right", padx=10, pady=5)
        self.txt_kakao = scrolledtext.ScrolledText(self.tab_kakao, bg="#ffffff", fg="#1e293b", font=("Malgun Gothic", 10), bd=0, spacing1=4, spacing3=4, insertbackground="#3a86ff")
        self.txt_kakao.pack(fill="both", expand=True, padx=0, pady=0)
        
        self.tab_place = tk.Frame(tab_control, bg="#1c2541")
        self.txt_place = scrolledtext.ScrolledText(self.tab_place, bg="#1c2541", fg="#ffffff", font=("Malgun Gothic", 10), bd=0)
        self.txt_place.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_blog = tk.Frame(tab_control, bg="#1c2541")
        self.txt_blog = scrolledtext.ScrolledText(self.tab_blog, bg="#1c2541", fg="#ffffff", font=("Malgun Gothic", 10), bd=0)
        self.txt_blog.pack(fill="both", expand=True, padx=5, pady=5)
        
        tab_control.add(self.tab_log, text=" 💻 구동로그 ")
        tab_control.add(self.tab_kakao, text=" 💬 카톡브리핑 ")
        tab_control.add(self.tab_place, text=" 📍 플레이스 ")
        tab_control.add(self.tab_blog, text=" ✍️ 블로그 ")

        self.tab_history = tk.Frame(tab_control, bg="#0f172a")
        _hist_ctrl = tk.Frame(self.tab_history, bg="#243256", padx=12, pady=8)
        _hist_ctrl.pack(fill="x", side="top")
        tk.Label(_hist_ctrl, text="업체:", font=("Malgun Gothic", 10, "bold"), fg="#cbd5e1", bg="#243256").pack(side="left")
        self.cmb_hist_store = ttk.Combobox(_hist_ctrl, width=18, state="readonly", font=("Malgun Gothic", 10), style="Hist.TCombobox")
        self.cmb_hist_store.pack(side="left", padx=(4, 12))
        tk.Label(_hist_ctrl, text="모드:", font=("Malgun Gothic", 10, "bold"), fg="#cbd5e1", bg="#243256").pack(side="left")
        self.cmb_hist_mode = ttk.Combobox(_hist_ctrl, values=["플레이스", "블로그"], width=8, state="readonly", font=("Malgun Gothic", 10), style="Hist.TCombobox")
        self.cmb_hist_mode.set("플레이스")
        self.cmb_hist_mode.pack(side="left", padx=(4, 12))
        tk.Button(_hist_ctrl, text="🔄 새로고침", font=("Malgun Gothic", 8, "bold"), bg="#334155", fg="#f8fafc", bd=0, padx=7, pady=5, cursor="hand2", command=self.refresh_history_tab).pack(side="left", padx=(0, 6))
        tk.Button(_hist_ctrl, text="📥 Excel 저장", font=("Malgun Gothic", 8, "bold"), bg="#10b981", fg="white", bd=0, padx=7, pady=5, cursor="hand2", command=self.export_history_excel).pack(side="left")
        self.cmb_hist_store.bind("<<ComboboxSelected>>", lambda e: self.refresh_history_tab())
        self.cmb_hist_mode.bind("<<ComboboxSelected>>", lambda e: self.refresh_history_tab())

        _hist_nb = ttk.Notebook(self.tab_history)
        _hist_nb.pack(fill="both", expand=True)

        # ── Sub-tab 1: 순위표 ──
        _tab_table = tk.Frame(_hist_nb, bg="#0f172a")
        _hist_legend = tk.Frame(_tab_table, bg="#0f172a", padx=10)
        _hist_legend.pack(fill="x", side="top", pady=(4, 2))
        for _txt, _col in [("↑ 상승", "#10b981"), ("↓ 하락", "#ef4444"), ("→ 유지", "#3a86ff"), ("- 미노출", "#64748b"), ("⚠ 미검색", "#f59e0b")]:
            tk.Label(_hist_legend, text=_txt, font=("Malgun Gothic", 8), fg=_col, bg="#0f172a").pack(side="left", padx=8)
        _hist_txt_frame = tk.Frame(_tab_table, bg="#0f172a")
        _hist_txt_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))
        _hist_xsb = tk.Scrollbar(_hist_txt_frame, orient="horizontal")
        _hist_xsb.pack(side="bottom", fill="x")
        _hist_ysb = tk.Scrollbar(_hist_txt_frame, orient="vertical")
        _hist_ysb.pack(side="right", fill="y")
        self.txt_history = tk.Text(_hist_txt_frame, bg="#0f172a", fg="#f8fafc", font=("Consolas", 9), bd=0, wrap="none",
                                   xscrollcommand=_hist_xsb.set, yscrollcommand=_hist_ysb.set)
        self.txt_history.pack(side="left", fill="both", expand=True)
        _hist_xsb.config(command=self.txt_history.xview)
        _hist_ysb.config(command=self.txt_history.yview)
        for _tag, _col in [("up", "#10b981"), ("down", "#ef4444"), ("same", "#f8fafc"), ("none", "#64748b"), ("header", "#f8fafc"), ("kw", "#cbd5e1"), ("custom_kw", "#fbbf24"), ("missing", "#f59e0b"), ("hist", "#f8fafc")]:
            self.txt_history.tag_config(_tag, foreground=_col)
        _hist_nb.add(_tab_table, text=" 📊 순위표 ")

        # ── Sub-tab 2: 트렌드 브리핑 ──
        self._hist_nb = _hist_nb
        self.tab_briefing = tk.Frame(_hist_nb, bg="#0f172a")
        _brief_bar = tk.Frame(self.tab_briefing, bg="#1c2541", pady=6)
        _brief_bar.pack(fill="x", side="top")
        tk.Label(_brief_bar, text="최근 검색결과 기준  ·  순위 추세 브리핑", font=("Malgun Gothic", 9), fg="#64748b", bg="#1c2541").pack(side="left", padx=14)
        tk.Button(_brief_bar, text="  📋 복사  ", font=("Malgun Gothic", 9, "bold"), bg="#3a86ff", fg="white", bd=0, padx=10, pady=3, cursor="hand2", relief="flat", command=self.copy_briefing_to_clipboard).pack(side="right", padx=10, pady=3)
        _brief_txt_frame = tk.Frame(self.tab_briefing, bg="#0f172a")
        _brief_txt_frame.pack(fill="both", expand=True, padx=5, pady=5)
        _brief_ysb = tk.Scrollbar(_brief_txt_frame, orient="vertical")
        _brief_ysb.pack(side="right", fill="y")
        self.txt_briefing = tk.Text(_brief_txt_frame, bg="#0b1329", fg="#f8fafc",
                                    font=("Malgun Gothic", 10), bd=0, wrap="word",
                                    yscrollcommand=_brief_ysb.set, padx=20, pady=16,
                                    tabs=(185,))
        self.txt_briefing.pack(side="left", fill="both", expand=True)
        _brief_ysb.config(command=self.txt_briefing.yview)
        for _bt, _bc, _bf in [
            ("b_title",  "#f8fafc",  ("Malgun Gothic", 13, "bold")),
            ("b_store",  "#3a86ff",  ("Malgun Gothic", 11, "bold")),
            ("b_date",   "#475569",  ("Malgun Gothic", 9)),
            ("b_sep",    "#1e293b",  ("Malgun Gothic", 6)),
            ("b_sup",    "#10b981",  ("Malgun Gothic", 10, "bold")),
            ("b_sok",    "#3a86ff",  ("Malgun Gothic", 10, "bold")),
            ("b_sdn",    "#ef4444",  ("Malgun Gothic", 10, "bold")),
            ("b_snew",   "#fbbf24",  ("Malgun Gothic", 10, "bold")),
            ("b_kw",     "#e2e8f0",  ("Malgun Gothic Semilight", 10)),
            ("b_rup",    "#10b981",  ("Malgun Gothic", 10, "bold")),
            ("b_rok",    "#94a3b8",  ("Malgun Gothic", 10)),
            ("b_rdn",    "#ef4444",  ("Malgun Gothic", 10, "bold")),
            ("b_rnew",   "#fbbf24",  ("Malgun Gothic", 10, "bold")),
            ("b_hist",   "#b0bec5",  ("Malgun Gothic", 9)),
            ("b_stat",   "#94a3b8",  ("Malgun Gothic", 9)),
            ("b_foot",   "#1e293b",  ("Malgun Gothic", 8)),
        ]:
            self.txt_briefing.tag_config(_bt, foreground=_bc, font=_bf)
        _hist_nb.add(self.tab_briefing, text=" 💬 트렌드 브리핑 ")
        _hist_nb.bind("<<NotebookTabChanged>>", lambda e: self.refresh_history_briefing() if _hist_nb.index("current") == 1 else None)

        tab_control.add(self.tab_history, text=" 📊 분석기록 ")

        if FIREBASE_AVAILABLE and self.auth and self.auth.is_admin:
            self.tab_admin = tk.Frame(tab_control, bg="#0b1329", padx=15, pady=12)
            AdminTab(self.tab_admin, self.auth)
            tab_control.add(self.tab_admin, text=" 🔧 관리자 ")

        tab_control.pack(fill="both", expand=True)
        self.tab_control = tab_control

        # 마우스휠 스크롤 — ttk.Notebook이 bind_all을 덮어쓰므로 모든 위젯 생성 후 마지막에 바인딩
        def _on_mousewheel(event):
            # Text/Listbox/Treeview는 자체 스크롤 처리, 그 외는 캔버스 스크롤
            if isinstance(event.widget, (tk.Text, tk.Listbox, ttk.Treeview)):
                return
            _canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.root.bind_all("<MouseWheel>", _on_mousewheel)

        # 시작 시 분석 기록 드롭다운 자동 로드
        self.root.after(200, self.refresh_history_tab)

    def blink_status(self):
        if not self.is_running:
            self.lbl_status_led.config(fg="#ef4444")
            self.lbl_status_text.config(text="엔진 대기 중 (정지됨)", fg="#94a3b8")
            return
        current_fg = self.lbl_status_led.cget("foreground")
        next_fg = "#10b981" if current_fg != "#10b981" else "#1c2541"
        self.lbl_status_led.config(fg=next_fg)
        
        mode_text = "플레이스" if self.analysis_mode == "place" else "블로그"
        eta_display = f" | {self.eta_string}" if self.eta_string else ""
        self.lbl_status_text.config(text=f"엔진 실시간 탐색 중... ({mode_text}){eta_display}")
        
        self.root.after(500, self.blink_status)
            
    def on_treeview_select(self, event):
        selected_items = self.tree.selection()
        if not selected_items: return
        item = selected_items[0]
        values = self.tree.item(item, "values")
        store_name = values[0]
        for store in self.config["stores"]:
            if store["store_name"] == store_name:
                self.ent_store_name.delete(0, tk.END); self.ent_store_name.insert(0, store["store_name"])
                self.ent_place_url.delete(0, tk.END); self.ent_place_url.insert(0, store["place_url"])
                self.txt_custom_kw.delete("1.0", tk.END)
                custom_kws = store.get("custom_keywords", [])
                if custom_kws:
                    self.txt_custom_kw.insert("1.0", ", ".join(custom_kws))
                break
            
    def refresh_store_table(self, select_name=None):
        self.config = normalize_config(self.config)
        for item in self.tree.get_children():
            self.tree.delete(item)
        selected_iid = None
        for store in self.config["stores"]:
            name = store.get("store_name", "")
            url = store.get("place_url", "")
            iid = self.tree.insert("", "end", values=(name, url))
            if select_name and name == select_name:
                selected_iid = iid
        if selected_iid:
            self.tree.selection_set(selected_iid)
            self.tree.focus(selected_iid)
            self.tree.see(selected_iid)
            
    def save_custom_keywords(self):
        name = self.ent_store_name.get().strip()
        if not name:
            messagebox.showwarning("업체 미선택", "먼저 목록에서 업체를 선택하세요.")
            return
        raw = self.txt_custom_kw.get("1.0", tk.END).strip()
        custom_kws = [k.strip() for k in re.split(r'[,\n]', raw) if k.strip()]
        self.config = normalize_config(self.config)
        for s in self.config["stores"]:
            if s.get("store_name") == name:
                s["custom_keywords"] = custom_kws
                break
        else:
            messagebox.showwarning("업체 미선택", f"[{name}] 업체를 목록에서 먼저 선택하세요.")
            return
        self.save_current_settings()
        self.log(f"✅ [{name}] 직접 키워드 {len(custom_kws)}개 저장 완료")

    def add_store(self):
        name = self.ent_store_name.get().strip()
        url = self.ent_place_url.get().strip()
        if not name or not url:
            messagebox.showwarning("입력 확인", "광고주 업체명과 플레이스 URL을 모두 입력해 주세요.")
            return
        if not ("naver.me" in url or "naver.com" in url):
            if not messagebox.askyesno("URL 확인", "네이버 플레이스 URL 형식이 아닌 것 같습니다. 그래도 저장할까요?"):
                return

        raw_kw_text = self.txt_custom_kw.get("1.0", tk.END).strip()
        custom_kws = [k.strip() for k in re.split(r'[,\n]', raw_kw_text) if k.strip()]
        try:
            self.config = normalize_config(self.config)
            # 같은 업체명은 새 URL로 덮어쓰기
            self.config["stores"] = [s for s in self.config["stores"] if s.get("store_name") != name]
            self.config["stores"].append({"store_name": name, "place_url": url, "custom_keywords": custom_kws})

            # 저장 실패가 나도 화면 목록에는 바로 보이도록 먼저 갱신
            self.refresh_store_table(select_name=name)
            self.save_current_settings()

            self.log(f"🟢 [성공] 광고주 [{name}]가 목록에 등록되었습니다.")
            self.log(f"    ㄴ 저장파일: {CONFIG_FILE}")
            self.ent_store_name.delete(0, tk.END)
            self.ent_place_url.delete(0, tk.END)
        except Exception as e:
            self.refresh_store_table(select_name=name)
            messagebox.showerror("신규 등록 오류", f"목록 반영 또는 저장 중 문제가 발생했습니다.\n\n{e}")
            self.log(f"🔴 [등록 오류] {e}")
        
    def delete_store(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showinfo("알림", "삭제할 광고주를 목록에서 선택해 주세요.")
            return
        self.config = normalize_config(self.config)
        selected_names = []
        for item in selected_item:
            values = self.tree.item(item, "values")
            if values:
                selected_names.append(values[0])
        self.config["stores"] = [s for s in self.config["stores"] if s.get("store_name") not in selected_names]
        self.refresh_store_table()
        self.save_current_settings()
        for n in selected_names:
            self.log(f"🗑️ [삭제] 광고주 [{n}]가 삭제되었습니다.")
        
    def save_current_settings(self):
        self.config = normalize_config(self.config)
        self.config["safe_mode"] = self.var_safe_mode.get()
        self.config["kakao_style"] = self.cmb_kakao_style.get()
        save_config(self.config)
        if self.auth:
            try:
                self.auth.save_user_stores(self.config.get('stores', []))
            except Exception:
                pass
        
    def log(self, message):
        self.txt_log.insert(tk.END, message + "\n")
        self.txt_log.see(tk.END)
        
    def copy_to_clipboard(self):
        text = self.txt_kakao.get("1.0", tk.END).strip()
        if not text: messagebox.showinfo("알림", "복사할 브리핑 결과가 없습니다."); return
        self.root.clipboard_clear(); self.root.clipboard_append(text)
        messagebox.showinfo("복사 완료", "카톡 브리핑이 클립보드에 완벽하게 복사되었습니다!")

    def copy_log_to_clipboard(self):
        text = self.txt_log.get("1.0", tk.END).strip()
        if not text: return
        self.root.clipboard_clear(); self.root.clipboard_append(text)
        messagebox.showinfo("복사 완료", "로그가 클립보드에 복사되었습니다.")

    def request_stop(self):
        self.stop_requested = True
        self.btn_stop.config(text="⏳ 중지 처리 중...", state="disabled", bg="#64748b")

    def start_analysis_place(self):
        self.analysis_mode = "place"; self.start_analysis_thread(force_all=self.var_bulk_place.get())

    def start_analysis_blog(self):
        self.analysis_mode = "blog"; self.start_analysis_thread(force_all=self.var_bulk_blog.get())

    def start_analysis_thread(self, force_all=False):
        try:
            if self.is_running: return
            target_stores = []
            if force_all:
                target_stores = self.config.get("stores", [])
                if not target_stores:
                    messagebox.showinfo("알림", "등록된 광고주가 없습니다. 먼저 신규 등록을 해주세요.")
                    return
            else:
                selected_items = self.tree.selection()
                if selected_items:
                    for item in selected_items:
                        values = self.tree.item(item, "values")
                        for s in self.config["stores"]:
                            if s["store_name"] == values[0]: target_stores.append(s); break
                else:
                    if not self.config.get("stores"):
                        messagebox.showinfo("알림", "등록된 광고주가 없습니다. 먼저 신규 등록을 해주세요.")
                        return
                    if messagebox.askyesno("전체 분석 확인", "표에서 마우스로 선택한 업체가 없습니다. 현재 등록된 [전체 업체] 분석을 가동할까요?"):
                        target_stores = self.config["stores"]
                    else: return

            self.save_current_settings()
            self.is_running = True
            self.stop_requested = False
            self.eta_string = "계산 중..."

            self.btn_place_run.config(bg="#1a5e45", fg="#6b9e87", state="normal")
            self.btn_blog_run.config(bg="#1a5e45", fg="#6b9e87", state="normal")
            self.btn_stop.config(text="🛑 분석 중지", state="normal", bg="#ef4444")
            
            self.tab_control.select(0)
            self.txt_log.delete("1.0", tk.END)
            self.txt_kakao.delete("1.0", tk.END)
            if self.analysis_mode == "place":
                self.txt_place.delete("1.0", tk.END)
            else:
                self.txt_blog.delete("1.0", tk.END)
            
            self.blink_status()
            threading.Thread(target=self.run_async_loop, args=(target_stores,), daemon=True).start()
        except Exception as e:
            messagebox.showerror("엔진 구동 에러", str(e))
            self.is_running = False
        
    def run_async_loop(self, target_stores):
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(self.run_crawler_automation(target_stores))
            loop.close()
        except Exception as e:
            import traceback
            error_msg = f"엔진 처리 중 오류가 발생했습니다:\n\n{str(e)}\n{traceback.format_exc()}"
            self.root.after(0, lambda: messagebox.showerror("시스템 탐색 오류", error_msg))
        finally:
            gc.collect()
            self.root.after(0, self.on_analysis_complete)
        
    def _parse_rank_num(self, rank_str):
        s = str(rank_str)
        if not s or s in ("분석안함", "분석안함", "-", "", "미검색") or "밖" in s:
            return 999
        m = re.search(r'(\d+)', s)
        if not m:
            return 999
        # "30위 밖" → 31로 처리 (실제 순위보다 나쁨을 반영)
        return int(m.group(1)) + 1 if "밖" in s else int(m.group(1))

    def _update_history_stores(self):
        hist = load_history()
        hist_names = set(hist.keys())
        # gui_config 등록 순서 우선, 기록 있는 업체만
        cfg_names = [s.get("store_name", "") for s in self.config.get("stores", [])
                     if s.get("store_name") in hist_names]
        # gui_config에 없지만 기록 있는 업체는 뒤에 추가
        extra = [n for n in hist.keys() if n not in cfg_names]
        names = cfg_names + extra
        self.cmb_hist_store["values"] = names
        cur = self.cmb_hist_store.get()
        if names and (not cur or cur not in names):
            self.cmb_hist_store.set(names[0])

    def refresh_history_tab(self):
        self._update_history_stores()
        store_name = self.cmb_hist_store.get()
        mode = "place" if self.cmb_hist_mode.get() == "플레이스" else "blog"
        self.txt_history.config(state="normal")
        self.txt_history.delete("1.0", tk.END)
        if not store_name:
            self.txt_history.insert(tk.END, "  업체를 선택하세요.\n")
            self.txt_history.config(state="disabled")
            return
        hist = load_history()
        entries = [e for e in hist.get(store_name, []) if e.get("mode") == mode]
        if not entries:
            self.txt_history.insert(tk.END, f"  [{store_name}] 저장된 {self.cmb_hist_mode.get()} 분석 기록이 없습니다.\n  분석을 먼저 실행해 주세요.\n")
            self.txt_history.config(state="disabled")
            return
        entries.sort(key=lambda e: e["date"])
        dates = [e["date"][5:] for e in entries]  # MM-DD
        all_kws = list(dict.fromkeys(r["keyword"] for e in entries for r in e["results"]))
        kw_rank_map = {}
        for e in entries:
            for r in e["results"]:
                kw_rank_map.setdefault(r["keyword"], {})[e["date"][5:]] = (r.get("rank", "-"), r.get("is_custom", False))

        # 일반 키워드 / 키워드추가 분리
        normal_kws = [kw for kw in all_kws if not any(v[1] for v in kw_rank_map.get(kw, {}).values())]
        custom_kws  = [kw for kw in all_kws if     any(v[1] for v in kw_rank_map.get(kw, {}).values())]

        # 일반 키워드: 전체 날짜 중 10위 이내 진입 이력 있는 것만 표시
        def _ever_top10(kw):
            for v in kw_rank_map.get(kw, {}).values():
                if v[0] and v[0] not in ("-", "", "분석안함") and self._parse_rank_num(v[0]) <= 10:
                    return True
            return False
        filtered_normal_kws = [kw for kw in normal_kws if _ever_top10(kw)]

        # 한국어 시각적 폭(2컬럼) 기반 정렬 헬퍼
        def _dw(s):
            return sum(2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1 for c in s)
        kw_disp = 30  # 키워드 컬럼 시각 폭 (표시 기준)
        def _pad_kw(s):
            dw = _dw(s)
            if dw >= kw_disp - 1:
                out, cur = '', 0
                for c in s:
                    cw = 2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1
                    if cur + cw > kw_disp - 2:
                        return out + '… '
                    out += c; cur += cw
                return out + ' ' * max(0, kw_disp - _dw(out))
            return s + ' ' * (kw_disp - dw)

        # ── 픽셀 기반 탭스톱: 한국어+ASCII 혼합 폰트에서 정확한 열 정렬 ──────
        _tf     = tkfont.Font(family="Consolas", size=9)
        _kw_px  = _tf.measure("가" * (kw_disp // 2)) + 14
        _dt_px  = max(_tf.measure(s) for s in ["30위 밖", "미검색", "20위"]) + 18
        _tab_str = str(_kw_px)
        for _i in range(len(dates)):
            _tab_str += f" {_kw_px + (_i + 1) * _dt_px} right"
        _tab_str += f" {_kw_px + len(dates) * _dt_px + 8}"
        self.txt_history.config(tabs=_tab_str)

        date_col = 7
        header = "키워드\t" + "\t".join(dates) + "\t\t추세\n"
        sep = "─" * (kw_disp + date_col * len(dates) + 6) + "\n"
        self.txt_history.insert(tk.END, header, "header")
        self.txt_history.insert(tk.END, sep, "header")

        def _render_row(kw):
            kw_data = kw_rank_map.get(kw, {})
            is_custom = any(v[1] for v in kw_data.values())
            kw_tag = "custom_kw" if is_custom else "kw"
            self.txt_history.insert(tk.END, _pad_kw(kw).rstrip() + "\t", kw_tag)

            def _norm(d):
                raw, _ = kw_data.get(d, ("-", False))
                if not raw or raw in ("-", "", "분석안함"):
                    return "30위 밖"
                return raw

            prev_num = None
            prev_is_out = True
            for d in dates:
                rank_str = _norm(d)
                rank_num = self._parse_rank_num(rank_str)
                _is_missing = rank_str == "미검색"
                _is_out = "밖" in rank_str
                _is_last = (d == dates[-1])
                if _is_out:
                    tag = "none"         # 30위 밖: 날짜 무관 회색
                elif _is_missing:
                    tag = "missing" if _is_last else "none"
                elif not _is_last:
                    tag = "hist"         # 이전 날짜 순위권: 흰색
                elif prev_is_out:
                    tag = "kw"           # 최근 날짜: 신규 진입 파랑
                elif rank_num < prev_num:
                    tag = "up"
                elif rank_num > prev_num:
                    tag = "down"
                else:
                    tag = "same"
                self.txt_history.insert(tk.END, rank_str + "\t", tag)
                prev_num = rank_num
                prev_is_out = _is_out or _is_missing

            if len(dates) >= 2:
                r1 = _norm(dates[-2])
                r2 = _norm(dates[-1])
                n1 = self._parse_rank_num(r1)
                n2 = self._parse_rank_num(r2)
                if r2 == "미검색":
                    trend, ttag = "⚠️", "missing"
                elif "밖" in r1 and "밖" in r2:
                    trend, ttag = "-", "none"
                elif n2 < n1:
                    trend, ttag = "↑", "up"
                elif n2 > n1:
                    trend, ttag = "↓", "down"
                else:
                    trend, ttag = "→", "same"
            else:
                trend, ttag = "-", "none"
            self.txt_history.insert(tk.END, "\t" + trend + "\n", ttag)

        # 신규 순위권 진입 분리 (이번 날짜에만 순위권, 이전은 전부 30위 밖)
        def _is_new_entry(kw):
            if len(dates) < 2:
                return False
            kw_data = kw_rank_map.get(kw, {})
            latest = kw_data.get(dates[-1], ("-", False))[0]
            if not latest or latest in ("-", "", "분석안함") or "밖" in latest or latest == "미검색":
                return False
            for d in dates[:-1]:
                prev = kw_data.get(d, ("-", False))[0]
                if prev and prev not in ("-", "", "분析안함") and "밖" not in prev and prev != "미검색":
                    return False
            return True

        new_entry_kws = [kw for kw in filtered_normal_kws if _is_new_entry(kw)]
        existing_kws  = [kw for kw in filtered_normal_kws if not _is_new_entry(kw)]

        for kw in existing_kws:
            _render_row(kw)

        if new_entry_kws:
            self.txt_history.insert(tk.END, sep, "header")
            self.txt_history.insert(tk.END, "  🆕 신규 순위권 진입\n", "kw")
            for kw in new_entry_kws:
                _render_row(kw)

        if custom_kws:
            self.txt_history.insert(tk.END, sep, "header")
            self.txt_history.insert(tk.END, "  키워드추가\n", "custom_kw")
            for kw in custom_kws:
                _render_row(kw)

        self.txt_history.config(state="disabled")
        self.refresh_history_briefing()


    def refresh_history_briefing(self):
        store_name = self.cmb_hist_store.get()
        mode = "place" if self.cmb_hist_mode.get() == "플레이스" else "blog"
        mode_label = "📍 플레이스" if mode == "place" else "✍️ 블로그"
        if not hasattr(self, "txt_briefing"):
            return
        self.txt_briefing.config(state="normal")
        self.txt_briefing.delete("1.0", tk.END)
        ins = self.txt_briefing.insert
        if not store_name:
            ins(tk.END, "  업체를 선택하세요.\n")
            self.txt_briefing.config(state="disabled")
            return
        hist = load_history()
        entries = sorted([e for e in hist.get(store_name, []) if e.get("mode") == mode], key=lambda e: e["date"])
        if not entries:
            ins(tk.END, f"  [{store_name}]\n  저장된 {self.cmb_hist_mode.get()} 분석 기록이 없습니다.\n  분석을 먼저 실행해 주세요.\n")
            self.txt_briefing.config(state="disabled")
            return
        dates = [e["date"][5:] for e in entries]
        all_kws = list(dict.fromkeys(r["keyword"] for e in entries for r in e["results"]))
        kw_rank_map = {}
        for e in entries:
            for r in e["results"]:
                kw_rank_map.setdefault(r["keyword"], {})[e["date"][5:]] = (r.get("rank", "-"), r.get("is_custom", False))

        def _pr(s):
            if not s or s in ("-", "", "30위 밖", "미검색", "분析안함", "분析안함", "분석안함"):
                return 999
            try:
                return int("".join(c for c in s if c.isdigit())) or 999
            except Exception:
                return 999

        def _rs(kw, d):
            return kw_rank_map.get(kw, {}).get(d, ("-", False))[0] or "-"

        def _latest(kw):
            for d in reversed(dates):
                r = _pr(_rs(kw, d))
                if r <= 20:
                    return r, _rs(kw, d)
            return 999, "30위 밖"

        def _trend(kw):
            valid = [_pr(_rs(kw, d)) for d in dates if _pr(_rs(kw, d)) <= 20]
            if not valid:
                return "out"
            if len(valid) == 1:
                return "new"
            if valid[-1] < valid[-2]:
                return "up"
            if valid[-1] > valid[-2]:
                return "down"
            return "same"

        # 키워드 분류
        normal_kws = [kw for kw in all_kws if not any(v[1] for v in kw_rank_map.get(kw, {}).values())]
        custom_kws  = [kw for kw in all_kws if     any(v[1] for v in kw_rank_map.get(kw, {}).values())]
        ranked   = [(kw, _latest(kw), _trend(kw)) for kw in normal_kws if _latest(kw)[0] <= 20]
        r_custom = [(kw, _latest(kw), _trend(kw)) for kw in custom_kws  if _latest(kw)[0] <= 20]

        new_kw   = sorted([x for x in ranked if x[2] == "new"], key=lambda x: x[1][0])
        main_kws = sorted([x for x in ranked if x[2] != "new"], key=lambda x: x[1][0])

        # 헤더
        sep  = "━" * 34
        thin = "─" * 34
        ins(tk.END, f"\n{sep}\n", "b_sep")
        ins(tk.END, f"{mode_label}  순위 브리핑\n", "b_title")
        ins(tk.END, f"{store_name}\n", "b_store")
        ins(tk.END, f"{thin}\n", "b_sep")
        ins(tk.END, "📅  " + "  ›  ".join(dates) + "\n\n", "b_date")

        # 행 렌더링: [키워드탭] [이전순위→이전순위→] [N위(현재) ↑↓]
        def _row(kw, rs, tr):
            rtag  = {"up": "b_rup", "down": "b_rdn", "same": "b_rok", "new": "b_rnew"}.get(tr, "b_rok")
            arrow = {"up": " ↑", "down": " ↓", "same": "", "new": " ★"}.get(tr, "")
            # 이전 순위들 (현재 제외)
            prev_parts = []
            for d in dates[:-1]:
                r = _rs(kw, d)
                prev_parts.append(r.replace("위", "") if _pr(r) <= 20 else "-")
            # 현재 순위
            cur_r = _rs(kw, dates[-1]) if dates else rs
            cur_display = (cur_r.replace("위", "") if _pr(cur_r) <= 20 else "-")
            hist_str = ("→".join(prev_parts) + "→") if prev_parts else ""
            ins(tk.END, f"  {kw}\t", "b_kw")
            if hist_str:
                ins(tk.END, hist_str, "b_hist")
            ins(tk.END, f"{cur_display}위(현재){arrow}\n", rtag)

        # 메인 키워드 (상승/유지/하락 통합, 현재순위 오름차순)
        for kw, (rn, rs), tr in main_kws:
            _row(kw, rs, tr)

        # 키워드추가
        if r_custom:
            ins(tk.END, f"\n{thin}\n", "b_sep")
            ins(tk.END, "  키워드추가\n", "b_snew")
            for kw, (rn, rs), tr in sorted(r_custom, key=lambda x: x[1][0]):
                _row(kw, rs, tr)

        # 신규 진입 (하단 별도)
        if new_kw:
            ins(tk.END, f"\n{thin}\n", "b_sep")
            ins(tk.END, "🆕  신규 진입\n", "b_snew")
            for kw, (rn, rs), tr in new_kw:
                _row(kw, rs, tr)

        # 요약
        rising_n  = len([x for x in main_kws if x[2] == "up"])
        stable_n  = len([x for x in main_kws if x[2] == "same"])
        falling_n = len([x for x in main_kws if x[2] == "down"])
        total = len(ranked) + len(r_custom)
        ins(tk.END, f"\n{sep}\n", "b_sep")
        ins(tk.END, f"총 {total}개 키워드 순위권", "b_stat")
        if main_kws:
            ins(tk.END, f"   상승 {rising_n}개 · 유지 {stable_n}개 · 하락 {falling_n}개", "b_stat")
            if new_kw:
                ins(tk.END, f" · 신규진입 {len(new_kw)}개", "b_rnew")
        ins(tk.END, "\n")
        ins(tk.END, "플레이스 마스터PRO · 실시간 검색 순위 모니터링\n", "b_foot")
        self.txt_briefing.config(state="disabled")

    def copy_briefing_to_clipboard(self):
        self.refresh_history_briefing()
        try:
            self.txt_briefing.config(state="normal")
            text = self.txt_briefing.get("1.0", tk.END).strip()
            self.txt_briefing.config(state="disabled")
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("복사 완료", "브리핑이 클립보드에 복사됐습니다.\n카카오톡에 붙여넣기 하세요.")
        except Exception as ex:
            messagebox.showerror("오류", str(ex))

    def export_history_excel(self):
        store_name = self.cmb_hist_store.get()
        mode = "place" if self.cmb_hist_mode.get() == "플레이스" else "blog"
        if not store_name:
            messagebox.showinfo("알림", "업체를 선택하세요.")
            return
        hist = load_history()
        entries = sorted([e for e in hist.get(store_name, []) if e.get("mode") == mode], key=lambda e: e["date"])
        if not entries:
            messagebox.showinfo("알림", "저장된 기록이 없습니다.")
            return
        dates = [e["date"] for e in entries]
        all_kws = list(dict.fromkeys(r["keyword"] for e in entries for r in e["results"]))
        kw_rank_map = {}
        for e in entries:
            for r in e["results"]:
                kw_rank_map.setdefault(r["keyword"], {})[e["date"]] = r.get("rank", "-")
        wb = Workbook()
        ws = wb.active
        ws.title = f"{store_name[:20]}_{self.cmb_hist_mode.get()}"
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.append(["키워드"] + dates)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1C2541")
            cell.font = Font(bold=True, color="F8FAFC")
            cell.alignment = Alignment(horizontal="center")
        green = PatternFill("solid", fgColor="D1FAE5")
        red   = PatternFill("solid", fgColor="FEE2E2")
        for kw in all_kws:
            row = [kw]
            for d in dates:
                row.append(kw_rank_map.get(kw, {}).get(d, "-"))
            ws.append(row)
            r_idx = ws.max_row
            prev = None
            for c_idx, d in enumerate(dates, start=2):
                cell = ws.cell(r_idx, c_idx)
                cell.alignment = Alignment(horizontal="center")
                cur = self._parse_rank_num(str(cell.value))
                if prev is not None and str(cell.value) not in ("-", "분석안함"):
                    if cur < prev: cell.fill = green
                    elif cur > prev: cell.fill = red
                if str(cell.value) not in ("-", "분석안함"):
                    prev = cur
        ws.column_dimensions["A"].width = 28
        for col in ws.iter_cols(min_col=2, max_col=len(dates)+1):
            for cell in col:
                ws.column_dimensions[cell.column_letter].width = 12
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', store_name)
        fname = os.path.join(BASE_DIR, f"기록_{safe_name}_{self.cmb_hist_mode.get()}.xlsx")
        wb.save(fname)
        messagebox.showinfo("Excel 저장 완료", f"저장 위치:\n{fname}")

    def on_analysis_complete(self):
        self.is_running = False
        self.eta_string = ""
        self.lbl_status_text.config(text="엔진 대기 중 (정지됨)", fg="#94a3b8")
        self.lbl_status_led.config(fg="#ef4444")
        self.btn_place_run.config(bg="#10b981", fg="white", state="normal")
        self.btn_blog_run.config(bg="#10b981", fg="white", state="normal")
        self.btn_stop.config(text="🛑 분석 중지", fg="white", state="disabled", bg="#64748b")
        self.tab_control.select(1)

    async def run_crawler_automation(self, target_stores):
        today = datetime.today().strftime('%Y-%m-%d')
        all_reports = []
        safe_mode = self.config.get("safe_mode", True)
        self.root.after(0, lambda: self.log("🚀 크롬 가상 엔진 기동 중 (Anti-Ban 세이프가드 가동)...\n"))

        N_PLACE = 5  # 플레이스 동시 처리 워커 수
        N_BLOG  = 3  # 블로그 동시 처리 워커 수 (페이지당 부하 큼)

        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox", "--blink-settings=imagesEnabled=false"]
                )
                context = await browser.new_context(
                    viewport={"width": 1280, "height": 800},
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                )
                # 업체 정보 조회 전용 페이지 (병렬 풀과 분리)
                detail_page = await context.new_page()
                await detail_page.add_init_script("delete navigator.__proto__.webdriver;")

                async def _new_pool_page():
                    pg = await context.new_page()
                    await pg.add_init_script("delete navigator.__proto__.webdriver;")
                    return pg

                for store in target_stores:
                    if self.stop_requested: break
                    store_name, place_url = store["store_name"], store["place_url"]
                    mode_text = "플레이스" if self.analysis_mode == "place" else "블로그"
                    self.root.after(0, lambda n=store_name, m=mode_text: self.log(f"📂 [{n}] {m} 분석 시작..."))
                    # 결과탭 업체 헤더 추가
                    from datetime import date as _date
                    _today = _date.today().strftime("%Y-%m-%d")
                    if self.analysis_mode == "place":
                        self.root.after(0, lambda n=store_name, d=_today: self.txt_place.insert(tk.END, f"\n{'─'*40}\n📂 {n}  ({d})\n{'─'*40}\n"))
                        self.root.after(0, lambda: self.txt_place.see(tk.END))
                    else:
                        self.root.after(0, lambda n=store_name, d=_today: self.txt_blog.insert(tk.END, f"\n{'─'*40}\n📂 {n}  ({d})\n{'─'*40}\n"))
                        self.root.after(0, lambda: self.txt_blog.see(tk.END))

                    place_id, addr, cat, menus, official_tags, nearby_station, kw_list = await get_place_details_failsafe(detail_page, place_url, lambda msg: self.root.after(0, lambda m=msg: self.log(m)))
                    if not place_id: continue
                    if self.stop_requested: break

                    _log_kw = lambda msg: self.root.after(0, lambda m=msg: self.log(m))
                    all_target_keywords = generate_highly_relevant_keywords(store_name, cat, addr, menus, official_tags, nearby_station, keyword_list=kw_list, log_func=_log_kw)

                    _recovery_kw_ranks = {}  # 사후 복구에서 채워짐 (kw → rank)

                    if self.analysis_mode == "blog":
                        _brand_base = re.sub(r"(본점|직영점|지점|점)$", "", store_name.strip()).strip()
                        _brand_parts = [bp for bp in re.split(r"\s+", _brand_base) if len(bp) >= 2]
                        _brand_only = set([store_name.strip(), _brand_base] + _brand_parts)
                        all_target_keywords = [kw for kw in all_target_keywords if kw not in _brand_only]
                        clean_menus = [re.sub(r'[^가-힣a-zA-Z0-9]', '', m) for m in menus if len(re.sub(r'[^가-힣a-zA-Z0-9]', '', m)) >= 2]
                        menu_kws = [kw for kw in all_target_keywords if any(cm in kw for cm in clean_menus)]
                        all_target_keywords = list(dict.fromkeys(menu_kws + all_target_keywords))[:30]

                    kw_count = len(all_target_keywords)
                    n_workers = N_PLACE if self.analysis_mode == "place" else N_BLOG
                    base_per_kw = 12.0 if self.analysis_mode == "blog" and safe_mode else 8.0 if self.analysis_mode == "blog" else 2.0 if safe_mode else 1.2
                    total_est_seconds = int(kw_count * base_per_kw / n_workers)
                    mins, secs = divmod(total_est_seconds, 60)
                    time_str = f"{mins}분 {secs}초" if mins > 0 else f"{secs}초"
                    self.eta_string = f"예상 완료: 약 {time_str} 후"

                    self.root.after(0, lambda c=kw_count, t=time_str, w=n_workers: self.log(f"    ㄴ 🔥 완벽 복구된 AI가 타겟 키워드 {c}개를 기획했습니다! (병렬 {w}개 동시처리, 예상 시간: 약 {t})"))
                    self.root.after(0, lambda: self.log(f"    👉 상위 30위 노출 결과 뽑아냅니다...\n"))

                    report_data = {"store_name": store_name, "date": today, "results": [], "mode": self.analysis_mode}
                    blog_url_cache = {}

                    # ── 페이지 풀 생성 (Queue = 자연스러운 동시성 제한) ─────────────
                    search_pool = asyncio.Queue()
                    insp_pool   = asyncio.Queue()
                    for _ in range(n_workers):
                        await search_pool.put(await _new_pool_page())
                    if self.analysis_mode == "blog":
                        for _ in range(n_workers):
                            await insp_pool.put(await _new_pool_page())

                    # ── 플레이스 키워드 태스크 ──────────────────────────────────────
                    async def _place_task(idx, keyword):
                        if self.stop_requested:
                            return (idx, keyword, None, False)
                        pg = await search_pool.get()
                        try:
                            encoded_kw = urllib.parse.quote(keyword)
                            self.root.after(0, lambda kw=keyword, i=idx+1, t=kw_count: self.log(f"  🔍 [{i}/{t}] '{kw}' 검색 중..."))
                            self.root.after(0, lambda kw=keyword, i=idx+1, t=kw_count: (self.txt_place.insert(tk.END, f"  🔍 [{i}/{t}] {kw} ...\n"), self.txt_place.see(tk.END)))
                            if safe_mode: await asyncio.sleep(random.uniform(0.05, 0.15))
                            place_rank, is_hit = None, False
                            try:
                                p_url = f"https://search.naver.com/search.naver?query={encoded_kw}&where=place&sm=tab_jum"
                                await pg.goto(p_url, wait_until="domcontentloaded", timeout=12000)
                                await pg.wait_for_timeout(900)
                                # place 링크 가진 li 5개 이상 로드 대기 (실제 결과카드 기준)
                                try:
                                    await pg.wait_for_function(
                                        """() => {
                                            let c = 0;
                                            for (const li of document.querySelectorAll('li')) {
                                                for (const a of li.querySelectorAll('a[href]')) {
                                                    const h = a.getAttribute('href') || '';
                                                    if (h.includes('map.naver') && /\\d{8,11}/.test(h)) { c++; break; }
                                                }
                                                if (c >= 5) return true;
                                            }
                                            return false;
                                        }""",
                                        timeout=5000
                                    )
                                except Exception: pass
                                # JS 직접 스크롤 (마우스 휠은 지도 영역에서 줌만 되고 스크롤 안됨)
                                _scroll_js = '''() => {
                                    let scrolled = 0;
                                    for (const c of document.querySelectorAll('ul, div')) {
                                        if (c.scrollHeight > c.clientHeight + 50 && c.clientHeight > 200) {
                                            const r = c.getBoundingClientRect();
                                            if (r.left < 800) { c.scrollTop = c.scrollHeight; scrolled++; }
                                        }
                                    }
                                    window.scrollTo(0, document.body.scrollHeight);
                                    return scrolled;
                                }'''
                                # 12회 스크롤 (병렬 처리 timing 마진 확보, rank 6+ 안정 로드)
                                for _ in range(12):
                                    await pg.evaluate(_scroll_js)
                                    await pg.wait_for_timeout(350)
                                _raw = await pg.evaluate('''() => {
                                    // 카드 단위 광고 필터 (광고+오가닉 동일 pid도 오가닉 정상 검출)
                                    const ranked_ids = [], all_ids = [];
                                    const seenR = new Set(), seenA = new Set();
                                    const pats = [
                                        /(?:place|restaurant|accommodation|hairshop|clinic|nail|gym|cafe|map|pcmap|store|outlink|entry\/place|local\/naver_place|pinId|beauty|spa|wellness)\\/(\\d{8,11})/i,
                                        /[?&](?:id|pid|placeId|bizId)=(\\d{8,11})/i
                                    ];
                                    const adRe = /\\/\\/(?:ader|ad)\\.(?:search\\.)?naver\\.com/i;
                                    const cardsInfo = new Map();
                                    for (const li of document.querySelectorAll('li')) {
                                        let pid = null, isAd = false;
                                        for (const a of li.querySelectorAll('a[href]')) {
                                            const h = a.getAttribute('href') || '';
                                            if (adRe.test(h)) isAd = true;
                                            if (!pid) {
                                                let d=''; try{d=decodeURIComponent(h);}catch(e){d=h;}
                                                for (const p of pats){const m=d.match(p);if(m){pid=m[1];break;}}
                                            }
                                        }
                                        if (!pid) {
                                            for (const el of li.querySelectorAll('[data-id],[data-place-id],[data-cid],[data-sid]')) {
                                                const p=el.getAttribute('data-id')||el.getAttribute('data-place-id')||el.getAttribute('data-cid')||el.getAttribute('data-sid');
                                                if(p&&/^\\d{8,11}$/.test(p)){pid=p;break;}
                                            }
                                        }
                                        if (!pid) continue;
                                        let parentReg=false; let anc=li.parentElement;
                                        while(anc){if(anc.tagName==='LI'&&cardsInfo.has(anc)){parentReg=true;break;}anc=anc.parentElement;}
                                        if (parentReg) continue;
                                        cardsInfo.set(li, {pid, isAd});
                                    }
                                    const sortedCards = Array.from(cardsInfo.entries()).sort(([a],[b])=>{
                                        const pos=a.compareDocumentPosition(b);
                                        if(pos&0x04)return -1; if(pos&0x02)return 1; return 0;
                                    });
                                    for (const [li, info] of sortedCards) {
                                        if (info.isAd) continue;
                                        if (seenR.has(info.pid)) continue;
                                        seenR.add(info.pid); ranked_ids.push(info.pid);
                                    }
                                    all_ids.push(...ranked_ids); seenR.forEach(x=>seenA.add(x));
                                    for (const [li, info] of sortedCards) {
                                        if (!seenA.has(info.pid)) { seenA.add(info.pid); all_ids.push(info.pid); }
                                    }
                                    try{const html=document.documentElement.innerHTML;const jp=/"(?:placeId|place_id|pid|sid|bizId)"\\s*:\\s*"?(\\d{8,11})"?/gi;let jm;
                                        while((jm=jp.exec(html))!==null){if(!seenA.has(jm[1])){seenA.add(jm[1]);all_ids.push(jm[1]);}}}catch(e){}
                                    return {ranked_ids, all_ids};
                                }''')
                                p_ids = _raw.get('ranked_ids', [])
                                p_ids_all = _raw.get('all_ids', [])
                                # 데스크탑 결과가 아예 없을 때만 모바일 재시도 (배포시점 로직)
                                if not p_ids:
                                    p_url_m = f"https://m.search.naver.com/search.naver?query={encoded_kw}&where=m_place"
                                    await pg.goto(p_url_m, wait_until="domcontentloaded", timeout=10000)
                                    await pg.wait_for_timeout(600)
                                    for _ in range(8):
                                        await pg.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                                        await pg.wait_for_timeout(250)
                                    m_ids = await pg.evaluate('''() => {
                                        const ids = [], seen = new Set();
                                        const pats = [
                                            /(?:place|restaurant|accommodation|hairshop|clinic|nail|gym|cafe|map|entry\/place|local\/naver_place)\\/(\\d{8,11})/i,
                                            /[?&](?:id|pid|placeId)=(\\d{8,11})/i
                                        ];
                                        for (const a of document.querySelectorAll("a[href]")) {
                                            try {
                                                const d = decodeURIComponent(a.getAttribute("href")||"");
                                                for (const p of pats) { const m=d.match(p); if(m&&!seen.has(m[1])){seen.add(m[1]);ids.push(m[1]);break;} }
                                            } catch(e) {}
                                        }
                                        for (const el of document.querySelectorAll("[data-id],[data-place-id],[data-cid],[data-sid]")) {
                                            const pid=el.getAttribute("data-id")||el.getAttribute("data-place-id")||el.getAttribute("data-cid")||el.getAttribute("data-sid");
                                            if(pid&&/^\\d{8,11}$/.test(pid)&&!seen.has(pid)){seen.add(pid);ids.push(pid);}
                                        }
                                        return ids;
                                    }''')
                                    self.root.after(0, lambda kw=keyword, n=len(m_ids), found=(place_id in m_ids): self.log(f"      ㄴ 모바일 p_ids={n}개, 감지={'O' if found else 'X'}"))
                                    if m_ids:
                                        p_ids = m_ids
                                # ── 추가 폴백: ranked_ids(href+data) 미포함 시 2페이지(start=16) 탐색 ──
                                if place_id and place_id not in p_ids:
                                    try:
                                        _p2_url = p_url + "&start=16"
                                        await pg.goto(_p2_url, wait_until="domcontentloaded", timeout=12000)
                                        await pg.wait_for_timeout(800)
                                        # JS 직접 스크롤 + 페이지2는 더 깊게 (rank 30까지 로드)
                                        for _ in range(10):
                                            await pg.evaluate(_scroll_js)
                                            await pg.wait_for_timeout(350)
                                        _p2_ids = await pg.evaluate('''() => {
                                            const ids = [], seen = new Set();
                                            const pats = [
                                                /(?:place|restaurant|accommodation|hairshop|clinic|nail|gym|cafe|map|pcmap|store|outlink|entry\\/place|local\\/naver_place|pinId|beauty|spa|wellness)\\/(\\d{8,11})/i,
                                                /[?&](?:id|pid|placeId|bizId)=(\\d{8,11})/i
                                            ];
                                            const adRe = /\\/\\/(?:ader|ad)\\.(?:search\\.)?naver\\.com/i;
                                            const cardsInfo = new Map();
                                            for (const li of document.querySelectorAll('li')) {
                                                let pid=null, isAd=false;
                                                for (const a of li.querySelectorAll('a[href]')) {
                                                    const h=a.getAttribute('href')||'';
                                                    if(adRe.test(h))isAd=true;
                                                    if(!pid){let d='';try{d=decodeURIComponent(h);}catch(e){d=h;}for(const p of pats){const m=d.match(p);if(m){pid=m[1];break;}}}
                                                }
                                                if(!pid){for(const el of li.querySelectorAll('[data-id],[data-place-id],[data-cid],[data-sid]')){const p=el.getAttribute('data-id')||el.getAttribute('data-place-id')||el.getAttribute('data-cid')||el.getAttribute('data-sid');if(p&&/^\\d{8,11}$/.test(p)){pid=p;break;}}}
                                                if(!pid)continue;
                                                let parentReg=false;let anc=li.parentElement;
                                                while(anc){if(anc.tagName==='LI'&&cardsInfo.has(anc)){parentReg=true;break;}anc=anc.parentElement;}
                                                if(parentReg)continue;
                                                cardsInfo.set(li,{pid,isAd});
                                            }
                                            const sorted=Array.from(cardsInfo.entries()).sort(([a],[b])=>{const pos=a.compareDocumentPosition(b);if(pos&0x04)return -1;if(pos&0x02)return 1;return 0;});
                                            for(const [li, info] of sorted){if(info.isAd)continue;if(seen.has(info.pid))continue;seen.add(info.pid);ids.push(info.pid);}
                                            return ids;
                                        }''')
                                        self.root.after(0, lambda kw=keyword, n=len(_p2_ids), found=(place_id in _p2_ids): self.log(
                                            f"      🔬 [{kw}] 2페이지 p_ids={n}개 감지={'O' if found else 'X'}"
                                        ))
                                        if _p2_ids and place_id in _p2_ids:
                                            _r2 = _p2_ids.index(place_id) + 16  # 2페이지 = rank 16+
                                            if _r2 <= 30:
                                                place_rank = f"{_r2}위"; is_hit = True
                                                self.root.after(0, lambda kw=keyword, r=_r2: self.log(f"      ↩️ {kw}: 2페이지 {r}위 검출"))
                                    except Exception as _fe:
                                        self.root.after(0, lambda kw=keyword, e=str(_fe): self.log(f"      ⚠️ [{kw}] 폴백오류: {e[:80]}"))
                                        pass
                                if not place_id:
                                    self.root.after(0, lambda kw=keyword: self.log(f"      ⚠️ [{kw}] place_id 없음 — 업체 URL 재등록 필요"))
                                elif place_id in p_ids:
                                    rank_num = p_ids.index(place_id) + 1
                                    self.root.after(0, lambda kw=keyword, r=rank_num, n=len(p_ids): self.log(
                                        f"      🔎 [{kw}] p_ids={n}개, 감지rank={r}" + (" → 30위 밖 처리" if r > 30 else "")
                                    ))
                                    if rank_num <= 30:
                                        place_rank = f"{rank_num}위"; is_hit = True
                                else:
                                    self.root.after(0, lambda kw=keyword, n=len(p_ids): self.log(
                                        f"      🔎 [{kw}] p_ids={n}개, place_id 미포함"
                                    ))
                                # pcmap 폴백 제거됨 (v8.36) — 거리 기반 결과라 키워드와 무관하게 매칭되는 부정확성 발생
                            except: pass
                            if is_hit:
                                self.root.after(0, lambda kw=keyword, pr=place_rank: (
                                    self.txt_place.insert(tk.END, f"  ✅ {kw} : {pr}\n"),
                                    self.txt_place.see(tk.END)
                                ))
                            return (idx, keyword, place_rank, is_hit)
                        finally:
                            await search_pool.put(pg)

                    # ── 블로그 키워드 태스크 ────────────────────────────────────────
                    async def _blog_task(idx, keyword):
                        if self.stop_requested:
                            return (idx, keyword, [])
                        pg = await search_pool.get()
                        ip = await insp_pool.get()
                        try:
                            self.root.after(0, lambda kw=keyword, i=idx+1, t=kw_count: self.log(f"  🔍 [{i}/{t}] '{kw}' 검색 중..."))
                            self.root.after(0, lambda kw=keyword, i=idx+1, t=kw_count: (self.txt_blog.insert(tk.END, f"  🔍 [{i}/{t}] {kw} ...\n"), self.txt_blog.see(tk.END)))
                            _log_fn = lambda msg: self.root.after(0, lambda m=msg: self.log(m))
                            blog_results = await check_blog_ranking_deep(
                                page=pg,
                                inspect_page=ip,
                                keyword=keyword,
                                store_name=store_name,
                                place_id=place_id,
                                address=addr,
                                log_func=_log_fn,
                                max_hits=5,
                                url_cache=blog_url_cache
                            )
                            return (idx, keyword, blog_results)
                        finally:
                            await search_pool.put(pg)
                            await insp_pool.put(ip)

                    # ── 병렬 실행 ───────────────────────────────────────────────────
                    if self.analysis_mode == "place":
                        tasks = [_place_task(i, kw) for i, kw in enumerate(all_target_keywords)]
                    else:
                        tasks = [_blog_task(i, kw) for i, kw in enumerate(all_target_keywords)]

                    raw_results = await asyncio.gather(*tasks)
                    raw_results.sort(key=lambda x: x[0])  # 키워드 원래 순서로 정렬

                    # ── 결과 처리 (순서 보장) ───────────────────────────────────────
                    already_seen_links = set()
                    for result in raw_results:
                        if self.analysis_mode == "place":
                            _, keyword, place_rank, is_hit = result
                            if is_hit:
                                report_data["results"].append({
                                    "keyword": keyword, "place_rank": place_rank,
                                    "blog_status": "분석안함", "blog_link": "",
                                    "blog_title": "", "blog_score": 0, "blog_reasons": ""
                                })
                                self.root.after(0, lambda kw=keyword, pr=place_rank: self.log(f"      🟢 [노출 발견!] '{kw}' ➡️ {pr}"))
                        else:
                            _, keyword, blog_results = result
                            for blog_res in blog_results:
                                b_status  = blog_res.get("status", "추적오류")
                                b_link    = blog_res.get("blog_link", "")
                                b_title   = blog_res.get("title", "")
                                b_score   = blog_res.get("score", 0)
                                b_reasons = ", ".join(blog_res.get("reasons", []))
                                if b_status in ["순위권 밖", "분석안함", "추적오류", "검색결과없음"]:
                                    continue
                                if b_link and b_link in already_seen_links:
                                    continue
                                already_seen_links.add(b_link)
                                report_data["results"].append({
                                    "keyword": keyword, "place_rank": "분석안함",
                                    "blog_status": b_status, "blog_link": b_link,
                                    "blog_title": b_title, "blog_score": b_score, "blog_reasons": b_reasons
                                })
                                self.root.after(0, lambda kw=keyword, b=b_status, t=b_title, l=b_link, s=b_score, r=b_reasons: self.log(f"      🟢 [블로그 노출 발견!] '{kw}' ➡️ {b}\n         제목: {t}\n         근거: {r} / 점수: {s}\n         🔗 {l}"))
                                self.root.after(0, lambda kw=keyword, b=b_status, t=b_title, l=b_link: (
                                    self.txt_blog.insert(tk.END, f"  ✅ {kw} : {b}\n     제목 : {t}\n     🔗 {l}\n\n"),
                                    self.txt_blog.see(tk.END)
                                ))

                    # ── 직접 키워드 분석 (30위 범위, 플레이스 모드만) ────────────────
                    c_raw = []
                    if self.analysis_mode == "place" and not self.stop_requested:
                        _custom_kws = [k for k in store.get("custom_keywords", []) if k.strip()]
                        if _custom_kws:
                            self.root.after(0, lambda: self.log(f"\n  ⭐ 직접 등록 키워드 분석 ({len(_custom_kws)}개) ─────"))
                            self.root.after(0, lambda: (self.txt_place.insert(tk.END, f"\n{'─'*30}\n"), self.txt_place.see(tk.END)))

                            async def _custom_place_task(c_idx, c_kw):
                                if self.stop_requested:
                                    return (c_idx, c_kw, "중지됨")
                                pg = await search_pool.get()
                                try:
                                    enc = urllib.parse.quote(c_kw)
                                    c_url = f"https://search.naver.com/search.naver?query={enc}&where=place&sm=tab_jum"
                                    _scroll_js = '''() => {
                                        let scrolled = 0;
                                        for (const c of document.querySelectorAll('ul, div')) {
                                            if (c.scrollHeight > c.clientHeight + 50 && c.clientHeight > 200) {
                                                const r = c.getBoundingClientRect();
                                                if (r.left < 800) { c.scrollTop = c.scrollHeight; scrolled++; }
                                            }
                                        }
                                        window.scrollTo(0, document.body.scrollHeight);
                                        return scrolled;
                                    }'''
                                    await pg.goto(c_url, wait_until="domcontentloaded", timeout=12000)
                                    await pg.wait_for_timeout(900)
                                    try:
                                        await pg.wait_for_function(
                                            """() => {
                                                let c = 0;
                                                for (const li of document.querySelectorAll('li')) {
                                                    for (const a of li.querySelectorAll('a[href]')) {
                                                        const h = a.getAttribute('href') || '';
                                                        if (h.includes('map.naver') && /\\d{8,11}/.test(h)) { c++; break; }
                                                    }
                                                    if (c >= 5) return true;
                                                }
                                                return false;
                                            }""",
                                            timeout=5000
                                        )
                                    except Exception: pass
                                    for _ in range(12):
                                        await pg.evaluate(_scroll_js)
                                        await pg.wait_for_timeout(350)
                                    c_ids = await pg.evaluate('''() => {
                                        const ids=[], seen=new Set();
                                        const pats=[/(?:place|restaurant|accommodation|hairshop|clinic|nail|gym|cafe|map|pcmap|store|outlink|entry\\/place|local\\/naver_place|pinId|beauty|spa|wellness)\\/(\\d{8,11})/i,/[?&](?:id|pid|placeId|bizId)=(\\d{8,11})/i];
                                        const adRe = /\\/\\/(?:ader|ad)\\.(?:search\\.)?naver\\.com/i;
                                        const cardsInfo = new Map();
                                        for (const li of document.querySelectorAll('li')) {
                                            let pid=null, isAd=false;
                                            for (const a of li.querySelectorAll('a[href]')) {
                                                const h=a.getAttribute('href')||'';
                                                if(adRe.test(h))isAd=true;
                                                if(!pid){let d='';try{d=decodeURIComponent(h);}catch(e){d=h;}for(const p of pats){const m=d.match(p);if(m){pid=m[1];break;}}}
                                            }
                                            if(!pid){for(const el of li.querySelectorAll('[data-id],[data-place-id],[data-cid],[data-sid]')){const p=el.getAttribute('data-id')||el.getAttribute('data-place-id')||el.getAttribute('data-cid')||el.getAttribute('data-sid');if(p&&/^\\d{8,11}$/.test(p)){pid=p;break;}}}
                                            if(!pid)continue;
                                            let parentReg=false;let anc=li.parentElement;
                                            while(anc){if(anc.tagName==='LI'&&cardsInfo.has(anc)){parentReg=true;break;}anc=anc.parentElement;}
                                            if(parentReg)continue;
                                            cardsInfo.set(li,{pid,isAd});
                                        }
                                        const sorted=Array.from(cardsInfo.entries()).sort(([a],[b])=>{const pos=a.compareDocumentPosition(b);if(pos&0x04)return -1;if(pos&0x02)return 1;return 0;});
                                        for(const [li,info] of sorted){if(info.isAd)continue;if(seen.has(info.pid))continue;seen.add(info.pid);ids.push(info.pid);}
                                        return ids;
                                    }''')
                                    # ── 직접키워드 2페이지(start=16) 폴백 ──────────────────────────
                                    if place_id not in c_ids:
                                        try:
                                            _cp2_url = c_url + "&start=16"
                                            await pg.goto(_cp2_url, wait_until="domcontentloaded", timeout=12000)
                                            await pg.wait_for_timeout(800)
                                            for _ in range(10):
                                                await pg.evaluate(_scroll_js)
                                                await pg.wait_for_timeout(350)
                                            _cp2_ids = await pg.evaluate('''() => {
                                                const ids=[],seen=new Set();
                                                const pats=[/(?:place|restaurant|accommodation|hairshop|clinic|nail|gym|cafe|map|pcmap|store|outlink|entry\\/place|local\\/naver_place|pinId|beauty|spa|wellness)\\/(\\d{8,11})/i,/[?&](?:id|pid|placeId|bizId)=(\\d{8,11})/i];
                                                const adRe = /\\/\\/(?:ader|ad)\\.(?:search\\.)?naver\\.com/i;
                                                const cardsInfo = new Map();
                                                for (const li of document.querySelectorAll('li')) {
                                                    let pid=null, isAd=false;
                                                    for (const a of li.querySelectorAll('a[href]')) {
                                                        const h=a.getAttribute('href')||'';
                                                        if(adRe.test(h))isAd=true;
                                                        if(!pid){let d='';try{d=decodeURIComponent(h);}catch(e){d=h;}for(const p of pats){const m=d.match(p);if(m){pid=m[1];break;}}}
                                                    }
                                                    if(!pid){for(const el of li.querySelectorAll('[data-id],[data-place-id],[data-cid],[data-sid]')){const p=el.getAttribute('data-id')||el.getAttribute('data-place-id')||el.getAttribute('data-cid')||el.getAttribute('data-sid');if(p&&/^\\d{8,11}$/.test(p)){pid=p;break;}}}
                                                    if(!pid)continue;
                                                    let parentReg=false;let anc=li.parentElement;
                                                    while(anc){if(anc.tagName==='LI'&&cardsInfo.has(anc)){parentReg=true;break;}anc=anc.parentElement;}
                                                    if(parentReg)continue;
                                                    cardsInfo.set(li,{pid,isAd});
                                                }
                                                const sorted=Array.from(cardsInfo.entries()).sort(([a],[b])=>{const pos=a.compareDocumentPosition(b);if(pos&0x04)return -1;if(pos&0x02)return 1;return 0;});
                                                for(const [li,info] of sorted){if(info.isAd)continue;if(seen.has(info.pid))continue;seen.add(info.pid);ids.push(info.pid);}
                                                return ids;
                                            }''')
                                            if _cp2_ids and place_id in _cp2_ids:
                                                _cr2 = _cp2_ids.index(place_id) + 16
                                                if _cr2 <= 30:
                                                    return (c_idx, c_kw, f"{_cr2}위")
                                        except Exception:
                                            pass
                                    # ── 모바일 최후 폴백 ────────────────────────────────────────────
                                    if place_id not in c_ids:
                                        m_url = f"https://m.search.naver.com/search.naver?query={enc}&where=m_place"
                                        await pg.goto(m_url, wait_until="domcontentloaded", timeout=10000)
                                        await pg.wait_for_timeout(600)
                                        for _ in range(8):
                                            await pg.mouse.wheel(0, 800)
                                            await pg.wait_for_timeout(250)
                                        m_c_ids = await pg.evaluate('''() => {const ids=[],seen=new Set();const pats=[/(?:place|restaurant|accommodation|hairshop|clinic|nail|gym|cafe|map|entry\\/place|local\\/naver_place)\\/(\\d{8,11})/i,/[?&](?:id|pid|placeId)=(\\d{8,11})/i];for(const a of document.querySelectorAll("a[href]")){try{const d=decodeURIComponent(a.getAttribute("href")||"");for(const p of pats){const m=d.match(p);if(m&&!seen.has(m[1])){seen.add(m[1]);ids.push(m[1]);break;}}}catch(e){}}for(const el of document.querySelectorAll("[data-id],[data-place-id],[data-cid],[data-sid]")){const pid=el.getAttribute("data-id")||el.getAttribute("data-place-id")||el.getAttribute("data-cid")||el.getAttribute("data-sid");if(pid&&/^\\d{8,11}$/.test(pid)&&!seen.has(pid)){seen.add(pid);ids.push(pid);}}return ids;}''')
                                        if m_c_ids: c_ids = m_c_ids
                                    if place_id in c_ids:
                                        r_num = c_ids.index(place_id) + 1
                                        if r_num <= 30:
                                            return (c_idx, c_kw, f"{r_num}위")
                                    return (c_idx, c_kw, "30위 밖")
                                except Exception:
                                    return (c_idx, c_kw, "30위 밖")
                                finally:
                                    await search_pool.put(pg)

                            c_tasks = [_custom_place_task(i, kw) for i, kw in enumerate(_custom_kws)]
                            c_raw = await asyncio.gather(*c_tasks)
                            c_raw.sort(key=lambda x: x[0])
                            for _, c_kw, c_rank in c_raw:
                                report_data["results"].append({
                                    "keyword": c_kw, "place_rank": c_rank,
                                    "blog_status": "분석안함", "blog_link": "",
                                    "blog_title": "", "blog_score": 0, "blog_reasons": "",
                                    "is_custom": True
                                })
                                _icon = "✅" if c_rank != "30위 밖" else "⬜"
                                self.root.after(0, lambda k=c_kw, r=c_rank, ic=_icon: (
                                    self.txt_place.insert(tk.END, f"  {ic} {k} : {r}\n"),
                                    self.txt_place.see(tk.END)
                                ))

                    # ── 사후 복구: 이전 순위권 키워드 중 이번 분석에서 빠진 것 추가 검색 ──
                    if not self.stop_requested:
                        _hist_rc = load_history()
                        _rc_mode = "place" if self.analysis_mode == "place" else "blog"
                        _rc_entries = [e for e in _hist_rc.get(store_name, []) if e.get("mode") == _rc_mode]
                        _searched_set = {kw for _, kw, *_ in raw_results} | {ckw for _, ckw, _ in c_raw}
                        _rc_needed = []
                        _rc_seen = set()
                        for _rce in _rc_entries:
                            for _rcp in _rce.get("results", []):
                                _rck = _rcp.get("keyword", "")
                                _rcr = _rcp.get("rank", "")
                                if (_rck
                                        and not _rcp.get("is_custom", False)
                                        and _rck not in _searched_set
                                        and _rck not in _rc_seen
                                        and _rcr not in ("30위 밖", "-", "", "분석안함", "분석안함", "분석안함", "미검색")
                                        and self._parse_rank_num(_rcr) <= 20):
                                    _rc_needed.append(_rck)
                                    _rc_seen.add(_rck)
                        if _rc_needed:
                            _n_rc = len(_rc_needed)
                            _orig_kc = kw_count
                            kw_count = _orig_kc + _n_rc
                            _kl_preview = ', '.join(_rc_needed[:3]) + (' 외' if _n_rc > 3 else '')
                            self.root.after(0, lambda n=_n_rc, p=_kl_preview: self.log(f"\n  🔄 사후 복구: 이전 순위권 키워드 {n}개 추가 분석 ({p})"))
                            if self.analysis_mode == "place":
                                _rc_tasks = [_place_task(_orig_kc + _ri, _rk) for _ri, _rk in enumerate(_rc_needed)]
                                _rc_raw = await asyncio.gather(*_rc_tasks)
                                for _, _rk, _rp, _rh in _rc_raw:
                                    _recovery_kw_ranks[_rk] = _rp if _rh else "30위 밖"
                                    if _rh:
                                        report_data["results"].append({
                                            "keyword": _rk, "place_rank": _rp,
                                            "blog_status": "분석안함", "blog_link": "",
                                            "blog_title": "", "blog_score": 0, "blog_reasons": ""
                                        })
                                        self.root.after(0, lambda kw=_rk, pr=_rp: (
                                            self.txt_place.insert(tk.END, f"  ✅ {kw} : {pr} [복구]\n"),
                                            self.txt_place.see(tk.END)
                                        ))
                            else:
                                _rc_tasks = [_blog_task(_orig_kc + _ri, _rk) for _ri, _rk in enumerate(_rc_needed)]
                                _rc_raw = await asyncio.gather(*_rc_tasks)
                                for _, _rk, _rbl in _rc_raw:
                                    _had_rc_hit = False
                                    for _rb in _rbl:
                                        _rb_st = _rb.get("status", "추적오류")
                                        _rb_lk = _rb.get("blog_link", "")
                                        if _rb_st in ["순위권 밖", "분석안함", "추적오류", "검색결과없음"]:
                                            continue
                                        if _rb_lk and _rb_lk in already_seen_links:
                                            continue
                                        already_seen_links.add(_rb_lk)
                                        _had_rc_hit = True
                                        _recovery_kw_ranks[_rk] = _rb_st
                                        report_data["results"].append({
                                            "keyword": _rk, "place_rank": "분석안함",
                                            "blog_status": _rb_st, "blog_link": _rb_lk,
                                            "blog_title": _rb.get("title", ""), "blog_score": _rb.get("score", 0),
                                            "blog_reasons": ", ".join(_rb.get("reasons", []))
                                        })
                                    if not _had_rc_hit:
                                        _recovery_kw_ranks[_rk] = "30위 밖"
                            kw_count = _orig_kc

                    # ── 기록 저장 ─────────────────────────────────────────────────────
                    try:
                        _hist = load_history()
                        _store_hist = _hist.setdefault(store_name, [])
                        _hist_mode = "place" if self.analysis_mode == "place" else "blog"
                        _hist_results = []
                        if _hist_mode == "place":
                            # 일반 키워드 전체 저장 (미검출 → "30위 밖")
                            for _, _kw, _pr, _hit in raw_results:
                                _hist_results.append({"keyword": _kw, "rank": _pr if _hit else "30위 밖", "is_custom": False})
                            # 직접 등록 키워드 전체 저장
                            for _, _ckw, _cr in c_raw:
                                _hist_results.append({"keyword": _ckw, "rank": _cr, "is_custom": True})
                        else:
                            for r in report_data["results"]:
                                _hist_results.append({"keyword": r["keyword"], "rank": r.get("blog_status", ""), "is_custom": r.get("is_custom", False)})
                        # 복구 분석 결과 추가 (추가 검색했으나 미검출 → "30위 밖", 검출 → 실제 순위)
                        _hist_rc_kw_set = {r["keyword"] for r in _hist_results if not r.get("is_custom", False)}
                        for _rk2, _rr2 in _recovery_kw_ranks.items():
                            if _rk2 not in _hist_rc_kw_set:
                                _hist_results.append({"keyword": _rk2, "rank": _rr2, "is_custom": False})
                        # 이전 기록에서 한번이라도 순위권에 진입한 키워드가
                        # 이번 실행의 키워드 목록에서 빠진 경우 "미검색"으로 보존
                        # (30위 밖 = 검색했으나 미검출 / 미검색 = 아예 검색 목록에서 제외됨)
                        _curr_kws = {r["keyword"] for r in _hist_results if not r.get("is_custom", False)}
                        for _pe in [e for e in _store_hist if e.get("mode") == _hist_mode]:
                            for _pr2 in _pe.get("results", []):
                                _pk = _pr2.get("keyword", "")
                                _prank = _pr2.get("rank", "")
                                if (_pk and not _pr2.get("is_custom", False)
                                        and _pk not in _curr_kws
                                        and _prank not in ("30위 밖", "-", "", "분석안함", "분석안함", "분석안함", "미검색")):
                                    _hist_results.append({"keyword": _pk, "rank": "미검색", "is_custom": False})
                                    _curr_kws.add(_pk)
                        _hist_entry = {
                            "date": today,
                            "mode": _hist_mode,
                            "results": _hist_results
                        }
                        # 같은 날짜+모드 기록은 덮어쓰기
                        _store_hist[:] = [e for e in _store_hist if not (e["date"] == today and e["mode"] == _hist_mode)]
                        _store_hist.append(_hist_entry)
                        # 최근 60일치만 유지
                        _store_hist.sort(key=lambda e: e["date"])
                        _hist[store_name] = _store_hist[-60:]
                        save_history(_hist)
                        # Firebase에 해당 업체 기록 백업
                        if self.auth:
                            try:
                                self.auth.save_user_history_store(store_name, _hist[store_name])
                            except Exception:
                                pass
                        self.root.after(0, lambda: self._update_history_stores())
                    except Exception:
                        pass

                    # ── 풀 페이지 정리 (예외가 report 수집을 막지 않도록 try/except) ─────
                    try:
                        while not search_pool.empty():
                            pg = await search_pool.get()
                            try: await pg.close()
                            except Exception: pass
                    except Exception: pass
                    if self.analysis_mode == "blog":
                        try:
                            while not insp_pool.empty():
                                pg = await insp_pool.get()
                                try: await pg.close()
                                except Exception: pass
                        except Exception: pass

                    n_results = len(report_data["results"])
                    self.root.after(0, lambda n=n_results: self.log(f"      📊 수집 완료: {n}개 결과\n" if n else "      ⚠️ 상위 10위 내에 진입한 결과가 없습니다.\n"))
                    all_reports.append(report_data)

                try:
                    await browser.close()
                except Exception:
                    pass

                style_mode = self.config.get("kakao_style", "이모지 Trendy형")
                final_briefing = []
                for r in all_reports:
                    try:
                        final_briefing.append(self.generate_report(r, style_mode))
                        final_briefing.append("\n" + "="*45 + "\n")
                    except Exception as ge:
                        self.root.after(0, lambda msg=str(ge): self.log(f"      ⚠️ 보고서 생성 오류: {msg}"))
                        final_briefing.append(f"[보고서 생성 오류: {ge}]\n")
                        final_briefing.append("\n" + "="*45 + "\n")

                self.root.after(0, lambda: self.txt_kakao.insert("1.0", "".join(final_briefing)))
        except Exception as e:
            self.root.after(0, lambda msg=str(e): messagebox.showerror("시스템 탐색 오류", f"구동 중 문제가 발생했습니다.\n\n{msg}"))

    def generate_report(self, report, style_mode):
        if not report: return ""

        _mode = report.get("mode", self.analysis_mode)  # report에 저장된 모드 우선 사용

        # 플레이스 모드: 띄어쓰기만 다른 동일 순위 키워드 중복 제거 (앞쪽 결과 우선 유지)
        if _mode == "place" and report.get("results"):
            seen_kw_rank = {}
            deduped = []
            for res in report["results"]:
                norm_kw = res.get("keyword", "").replace(" ", "")
                rank = res.get("place_rank", "")
                key = (norm_kw, rank)
                if key not in seen_kw_rank:
                    seen_kw_rank[key] = True
                    deduped.append(res)
            report = dict(report)
            report["results"] = deduped

        msg = []
        name, date = report["store_name"], report["date"]
        mode_str = "플레이스" if _mode == "place" else "블로그"
        is_blog = _mode == "blog"

        if not report["results"]:
            if style_mode == "비즈니스 격식형":
                msg.append(f"[플레이스 마스터PRO] {mode_str} 순위 보고서")
                msg.append(f"분석 일자 : {date} | 분석 대상 : {name}\n")
                msg.append("현재 상위 10위 내 진입한 키워드가 없습니다.")
                msg.append("지속적인 모니터링을 통해 순위 변화를 추적하겠습니다.")
            elif style_mode == "초간단 핵심요약형":
                msg.append(f"[{name}] {mode_str} 현황")
                msg.append(f"분석일 : {date}")
                msg.append("상위 노출 없음 — 지속 모니터링 중")
            else:
                msg.append(f"📊 [{name}] {mode_str} 순위 리포트\n")
                msg.append(f"📅 분석 일자 : {date}\n")
                msg.append("⚠️ 현재 상위에 진입한 키워드가 없습니다.\n지속적인 모니터링을 통해 순위 변화를 알려드리겠습니다.")
            return "\n".join(msg)

        _regular = [r for r in report["results"] if not r.get("is_custom")]
        _custom  = [r for r in report["results"] if r.get("is_custom")]

        if style_mode == "비즈니스 격식형":
            msg.append(f"[플레이스 마스터PRO] {mode_str} 순위 보고서")
            msg.append(f"분석 일자 : {date} | 분석 대상 : {name}\n")
            if is_blog:
                msg.append(f"[블로그 노출 현황]  (총 {len(_regular)}건)")
                for res in _regular:
                    msg.append(f"- {res['keyword']} : {res['blog_status']} 노출")
                    if res.get("blog_title"):
                        msg.append(f"  제목 : {res['blog_title']}")
                    if res.get("blog_link"):
                        msg.append(f"  링크 : {res['blog_link']}")
            else:
                msg.append("[플레이스 순위 현황]")
                for res in _regular:
                    msg.append(f"- {res['keyword']} : {res['place_rank']}")
                for res in _custom:
                    msg.append(f"- {res['keyword']} : {res['place_rank']}")
            msg.append("\n※ 플레이스 마스터PRO 자동 생성 보고서입니다.")

        elif style_mode == "초간단 핵심요약형":
            msg.append(f"[{name}] {mode_str} 현황  ({date})")
            if is_blog:
                for res in _regular:
                    msg.append(f"{res['keyword']} → {res['blog_status']} ✅")
                msg.append(f"총 {len(_regular)}건 노출 확인")
            else:
                for res in _regular:
                    rank_digits = re.findall(r'\d+', res["place_rank"])
                    rank_num = int(rank_digits[0]) if rank_digits else 99
                    icon = "👑" if rank_num == 1 else "⭐" if rank_num <= 3 else "🔥" if rank_num <= 10 else "▪"
                    msg.append(f"{icon} {res['keyword']} → {res['place_rank']}")
                for res in _custom:
                    msg.append(f"▪ {res['keyword']} → {res['place_rank']}")

        else:  # 이모지 Trendy형 (기본값)
            if is_blog:
                msg.append("안녕하세요 대표님^^")
                msg.append("오늘의 블로그포스팅 실시간 순위 리포트 보고드립니다.")
                msg.append(f"■일자 : {date} ")
                msg.append(f"■대상 : {name}\n")
                msg.append("📍 [블로그포스팅 실시간 순위 현황]")
                for res in _regular:
                    msg.append(f"  ✅ {res['keyword']} : {res['blog_status']}")
                    if res.get('blog_title'):
                        msg.append(f"     제목 : {res['blog_title']}")
                    if res.get('blog_link'):
                        msg.append(f"     🔗 {res['blog_link']}")
            else:
                msg.append("안녕하세요 대표님^^")
                msg.append("오늘의 플레이스 실시간 순위 리포트 보고드립니다.")
                msg.append(f"일자 : {date} ")
                msg.append(f"대상 : {name}\n")
                msg.append("📍 [플레이스 실시간 순위 현황]")
                for res in _regular:
                    rank_digits = re.findall(r'\d+', res["place_rank"])
                    rank_num = int(rank_digits[0]) if rank_digits else 99
                    icon = "👑" if rank_num == 1 else "⭐️" if rank_num <= 3 else "🔥" if rank_num <= 10 else "▪️"
                    msg.append(f"  {icon} {res['keyword']} : {res['place_rank']}")
                for res in _custom:
                    rank_digits = re.findall(r'\d+', res["place_rank"])
                    rank_num = int(rank_digits[0]) if rank_digits else 99
                    icon = "👑" if rank_num == 1 else "⭐️" if rank_num <= 3 else "🔥" if rank_num <= 10 else "🎯" if rank_num <= 30 else "⬜"
                    msg.append(f"  {icon} {res['keyword']} : {res['place_rank']}")
            msg.append("\n📊 오늘도 기분좋은 하루 보내세요^^")

        return "\n".join(msg)

if __name__ == "__main__":
    _script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))

    # firebase_config.json 있는데 인증모듈 로드 실패 → 배포판에서 직접 실행 차단
    _cfg_path = os.path.join(_script_dir, "firebase_config.json")
    if not FIREBASE_AVAILABLE and os.path.exists(_cfg_path):
        _r = tk.Tk(); _r.withdraw()
        import tkinter.messagebox as _mb
        _mb.showerror("실행 오류",
                      "플레이스마스터_실행.pyw 파일로 실행해주세요.\n\n"
                      "(_core.py 또는 naver_tracker.py 직접 실행 불가)")
        sys.exit(1)

    _auth = None
    if FIREBASE_AVAILABLE:
        _auth = load_firebase_auth()
        if _auth:
            if not try_auto_login(_auth, _script_dir):
                # 자동로그인 정보 없음 → 로그인창 표시
                _login = LoginWindow(_auth, script_dir=_script_dir)
                if not _login.run():
                    sys.exit(0)
            # 로그인 완료 (자동 또는 수동) → 업데이트 확인 후 진행
            check_and_update(_auth)

    root = tk.Tk()
    _ico = os.path.join(_script_dir, "icon.ico")
    if os.path.exists(_ico):
        try: root.iconbitmap(_ico)
        except Exception: pass
    app = PremiumMarketingApp(root, auth=_auth)
    root.mainloop()

